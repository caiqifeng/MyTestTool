"""
批量股票分析系统
功能：批量分析股票列表，生成综合Excel报告
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import time
import warnings
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
warnings.filterwarnings('ignore')

class BatchStockAnalyzer:
    def __init__(self, output_file='batch_stock_analysis.xlsx'):
        """
        初始化批量分析系统
        
        参数:
        output_file: 输出Excel文件名
        """
        self.output_file = output_file
        self.session = self._setup_session()
        
    def _setup_session(self):
        """设置请求会话"""
        session = requests.Session()
        session.trust_env = False  # 不读取系统代理
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        return session
    
    def get_stock_data_simple(self, stock_code):
        """
        简化版股票数据获取（多数据源尝试）
        返回：股票基本信息字典 或 None
        """
        code = str(stock_code).strip()
        
        # 数据源列表（按优先级排序）
        sources = [
            self._try_sina_source,
            self._try_tencent_source,
            self._try_eastmoney_simple,
        ]
        
        for source_func in sources:
            try:
                result = source_func(code)
                if result:
                    return result
            except:
                continue
            time.sleep(0.5)  # 请求间隔
        
        return None
    
    def _try_sina_source(self, code):
        """尝试新浪财经数据源"""
        try:
            market = "sh" if code.startswith('6') else "sz"
            url = f"http://hq.sinajs.cn/list={market}{code}"
            
            response = self.session.get(url, timeout=10)
            if response.status_code == 200:
                content = response.text
                
                # 解析新浪数据格式
                if "var hq_str_" in content:
                    data_str = content.split('"')[1] if '"' in content else ""
                    parts = data_str.split(',')
                    
                    if len(parts) >= 30:
                        # 新浪数据字段解析
                        stock_name = parts[0]
                        
                        # 尝试多个可能的当前价格字段
                        price_candidates = []
                        if len(parts) > 3 and parts[3]:  # 通常第4个字段是当前价
                            try:
                                price_candidates.append(float(parts[3]))
                            except:
                                pass
                        
                        if len(parts) > 1 and parts[1]:  # 开盘价
                            try:
                                price_candidates.append(float(parts[1]))
                            except:
                                pass
                        
                        current_price = price_candidates[0] if price_candidates else 0
                        
                        # 计算涨跌幅（如果有昨收价）
                        change_pct = 0
                        if len(parts) > 2 and parts[2] and current_price > 0:
                            try:
                                yesterday_close = float(parts[2])
                                if yesterday_close > 0:
                                    change_pct = ((current_price - yesterday_close) / yesterday_close) * 100
                            except:
                                pass
                        
                        return {
                            'code': code,
                            'name': stock_name,
                            'price': round(current_price, 2),
                            'change_pct': round(change_pct, 2),
                            'volume': parts[8] if len(parts) > 8 else '0',
                            'amount': parts[9] if len(parts) > 9 else '0',
                            'source': 'sina',
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
        except Exception as e:
            pass
        return None
    
    def _try_tencent_source(self, code):
        """尝试腾讯财经数据源"""
        try:
            url = f"http://qt.gtimg.cn/q=s_{code}"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                content = response.text
                parts = content.split('~')
                
                if len(parts) > 40:
                    stock_name = parts[1]
                    current_price = float(parts[3]) if parts[3] else 0
                    change_pct = float(parts[5].rstrip('%')) if parts[5] else 0
                    
                    return {
                        'code': code,
                        'name': stock_name,
                        'price': round(current_price, 2),
                        'change_pct': round(change_pct, 2),
                        'volume': parts[6] if len(parts) > 6 else '0',
                        'amount': parts[7] if len(parts) > 7 else '0',
                        'source': 'tencent',
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
        except:
            pass
        return None
    
    def _try_eastmoney_simple(self, code):
        """尝试东方财富简化接口"""
        try:
            secid = f"1.{code}" if code.startswith('6') else f"0.{code}"
            url = f"http://push2.eastmoney.com/api/qt/stock/get?invt=2&fltt=2&fields=f58,f43,f169,f170,f47,f48&secid={secid}"
            
            response = self.session.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data.get('rc') == 0:
                    item = data['data']
                    
                    current_price = item.get('f43', 0) / 100 if item.get('f43') else 0
                    change_pct = item.get('f170', 0)
                    
                    return {
                        'code': code,
                        'name': item.get('f58', f"股票{code}"),
                        'price': round(current_price, 2),
                        'change_pct': round(change_pct, 2),
                        'volume': str(item.get('f47', 0)),
                        'amount': str(item.get('f48', 0)),
                        'source': 'eastmoney',
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
        except:
            pass
        return None

    def _get_financials_basic(self, code):
        """
        从东方财富获取基础财务指标（市盈率、负债率、现金流）
        返回字典，可能包含键: 'pe', 'leverage', 'free_cash_flow'
        """
        if not code:
            return {}

        try:
            # 东方财富财务数据接口（包含 PE、负债率等）
            secid = f"1.{code}" if code.startswith('6') else f"0.{code}"
            # 字段说明：f40=市盈率, f116=负债率, f117=ROE
            url = f"http://push2.eastmoney.com/api/qt/stock/get?invt=2&fltt=2&fields=f40,f116,f117,f49,f50&secid={secid}"
            
            response = self.session.get(url, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data.get('rc') == 0:
                    item = data.get('data', {})
                    
                    # 提取 PE（市盈率）
                    pe = self._safe_float(item.get('f40'))
                    
                    # 提取负债率（作为杠杆指标）
                    leverage = self._safe_float(item.get('f116'))
                    
                    # 提取 ROE（权益报酬率，作为盈利能力指标）
                    roe = self._safe_float(item.get('f117'))
                    
                    # 现金流数据通常需要单独查询，这里用 ROE 作为近似（高ROE通常意味着良好现金生成能力）
                    free_cash_flow = roe  # 简化处理：用 ROE% 代表现金流得分的输入
                    
                    result = {}
                    if pe is not None:
                        result['pe'] = pe
                    if leverage is not None:
                        result['leverage'] = leverage / 100.0  # 转换为 0-1 范围
                    if free_cash_flow is not None:
                        result['free_cash_flow'] = free_cash_flow
                    
                    return result
        except Exception as e:
            pass
        
        return {}

    def _safe_float(self, val):
        """安全地将值转换为 float，失败返回 None"""
        if val is None or val == '':
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    
    def analyze_single_stock(self, stock_data):
        """
        分析单个股票
        
        参数:
        stock_data: 股票数据字典
        
        返回:
        包含分析结果的字典
        """
        if not stock_data:
            return {
                '股票代码': 'N/A',
                '股票名称': 'N/A',
                '当前价格': 0,
                '涨跌幅(%)': 0,
                '数据源': 'N/A',
                '分析结果': '数据获取失败',
                '操作建议': '无法分析',
                '风险等级': '高',
                '分析时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        
        price = stock_data['price']
        change_pct = stock_data['change_pct']

        # 尝试获取基础财务数据（杠杆、现金流、估值），若不可用则使用回退值
        fin = self._get_financials_basic(stock_data.get('code'))

        leverage = fin.get('leverage')  # 债务/股本或负债率（可缺省）
        free_cash_flow = fin.get('free_cash_flow')  # 自由现金流（可缺省）
        pe = fin.get('pe')  # 市盈率（可缺省）
        
        # 基于多维度指标计算评分（杠杆/现金流/估值）
        def score_leverage(val):
            # 债务越低越好；如果未知返回中性分数50
            if val is None:
                return 50
            try:
                # 假设 val 表示债务/股本或负债率（0~1 或 百分比）
                v = float(val)
                if v > 5:  # 极高杠杆
                    return 10
                if v > 2:
                    return 25
                if v > 1:
                    return 40
                if v > 0.5:
                    return 60
                if v > 0.2:
                    return 80
                return 95
            except:
                return 50

        def score_cashflow(val):
            # 自由现金流越大越好（正数越高分数越高）
            if val is None:
                return 50
            try:
                v = float(val)
                # 缩放到0-100
                if v < 0:
                    return 20
                if v < 1_000_000:
                    return 40
                if v < 10_000_000:
                    return 60
                if v < 100_000_000:
                    return 80
                return 95
            except:
                return 50

        def score_valuation(val):
            # PE 过高风险较高；合理区间给高分
            if val is None:
                return 50
            try:
                v = float(val)
                if v <= 0:
                    return 50
                if v < 10:
                    return 95
                if v < 20:
                    return 80
                if v < 40:
                    return 60
                if v < 80:
                    return 35
                return 10
            except:
                return 50

        lev_s = score_leverage(leverage)
        cf_s = score_cashflow(free_cash_flow)
        pe_s = score_valuation(pe)

        # 权重分配：现金流40%，杠杆30%，估值30%
        total_score = round((cf_s * 0.4 + lev_s * 0.3 + pe_s * 0.3), 1)

        # 根据总分给出风险等级
        if total_score >= 75:
            risk = '低'
            suggestion = '财务健康，建议关注/持有'
        elif total_score >= 55:
            risk = '中低'
            suggestion = '财务尚可，建议观察'
        elif total_score >= 40:
            risk = '中高'
            suggestion = '财务存在风险，谨慎操作'
        elif total_score >= 20:
            risk = '高'
            suggestion = '财务较差，建议回避或慎重'
        else:
            risk = '极高'
            suggestion = '财务严重风险，强烈回避'

        # 生成分析说明
        analysis = (f"涨跌幅 {change_pct}%；杠杆得分 {lev_s}，现金流得分 {cf_s}，估值得分 {pe_s}；"
                    f"综合得分 {total_score}，风险等级 {risk}")
        
        return {
            '股票代码': stock_data['code'],
            '股票名称': stock_data['name'],
            '当前价格': price,
            '涨跌幅(%)': change_pct,
            '成交量': stock_data.get('volume', 'N/A'),
            '成交额': stock_data.get('amount', 'N/A'),
            '数据源': stock_data['source'],
            '分析结果': analysis,
            '操作建议': suggestion,
            '风险等级': risk,
            '分析时间': stock_data['timestamp']
        }
    
    def analyze_stocks_batch(self, stock_list, max_workers=5):
        """
        批量分析股票
        
        参数:
        stock_list: 股票代码列表，如 ['600519', '000001', '300750']
        max_workers: 最大并发线程数
        
        返回:
        分析结果的DataFrame
        """
        print(f"🚀 开始批量分析 {len(stock_list)} 只股票...")
        print("=" * 60)
        
        results = []
        failed_stocks = []
        
        # 使用线程池并发获取数据
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            future_to_stock = {
                executor.submit(self.get_stock_data_simple, stock): stock 
                for stock in stock_list
            }
            
            # 处理完成的任务
            for i, future in enumerate(as_completed(future_to_stock), 1):
                stock_code = future_to_stock[future]
                
                try:
                    # 获取股票数据
                    stock_data = future.result(timeout=15)
                    
                    if stock_data:
                        # 分析股票
                        analysis_result = self.analyze_single_stock(stock_data)
                        results.append(analysis_result)
                        
                        print(f"✅ [{i}/{len(stock_list)}] {stock_code}: {stock_data['name']} "
                              f"价格:{stock_data['price']} 涨跌:{stock_data['change_pct']}%")
                    else:
                        failed_stocks.append(stock_code)
                        print(f"❌ [{i}/{len(stock_list)}] {stock_code}: 数据获取失败")
                        
                except Exception as e:
                    failed_stocks.append(stock_code)
                    print(f"❌ [{i}/{len(stock_list)}] {stock_code}: 分析失败 - {str(e)[:50]}")
                
                # 进度显示
                progress = (i / len(stock_list)) * 100
                if i % 5 == 0 or i == len(stock_list):
                    print(f"📊 进度: {i}/{len(stock_list)} ({progress:.1f}%)")
        
        # 汇总结果
        if results:
            df_results = pd.DataFrame(results)
            
            # 按涨跌幅排序
            df_results = df_results.sort_values('涨跌幅(%)', ascending=False)
            
            # 添加综合统计
            total_stocks = len(stock_list)
            success_stocks = len(results)
            success_rate = (success_stocks / total_stocks) * 100
            
            print(f"\n{'='*60}")
            print(f"📈 批量分析完成!")
            print(f"总计分析: {total_stocks} 只股票")
            print(f"成功分析: {success_stocks} 只 (成功率: {success_rate:.1f}%)")
            
            if failed_stocks:
                print(f"分析失败: {', '.join(failed_stocks[:10])}" + 
                      (f" 等{len(failed_stocks)}只" if len(failed_stocks) > 10 else ""))
            
            return df_results
        else:
            print("❌ 所有股票分析都失败了")
            return pd.DataFrame()
    
    def _merge_with_existing_data(self, df_results):
        """
        合并新分析结果与现有数据
        策略：保留每天最后一次分析，根据分析时间去重
        
        参数:
        df_results: 新的分析结果DataFrame
        
        返回:
        合并后的DataFrame
        """
        try:
            # 检查文件是否存在
            if not os.path.exists(self.output_file):
                print(f"📝 新建分析结果文件: {self.output_file}")
                return df_results.copy()
            
            # 读取现有数据
            try:
                existing_df = pd.read_excel(self.output_file, sheet_name='股票分析详情')
                print(f"📚 找到现有数据: {len(existing_df)} 条记录")
            except Exception as e:
                print(f"⚠️ 无法读取现有数据，将创建新文件: {e}")
                return df_results.copy()
            
            # 合并新旧数据
            merged_df = pd.concat([existing_df, df_results], ignore_index=True)
            
            # 转换分析时间为日期（用于去重）
            if '分析时间' in merged_df.columns:
                merged_df['分析日期'] = pd.to_datetime(merged_df['分析时间']).dt.date
                
                # 按股票代码和分析日期分组，只保留每天最后一次（按时间排序）
                # 先确保分析时间是datetime类型
                merged_df['分析时间'] = pd.to_datetime(merged_df['分析时间'])
                
                # 按股票代码和日期分组，取每组中时间最晚的记录
                merged_df = merged_df.sort_values('分析时间', ascending=False)
                merged_df = merged_df.drop_duplicates(subset=['股票代码', '分析日期'], keep='first')
                
                # 按股票代码和分析时间重新排序
                merged_df = merged_df.sort_values(['股票代码', '分析时间'], ascending=[True, False])
                merged_df = merged_df.drop('分析日期', axis=1)
                
                print(f"✨ 合并后总记录数: {len(merged_df)} 条")
                print(f"   （去重后保留每天最后一次分析）")
                
                return merged_df.reset_index(drop=True)
            else:
                print(f"⚠️ 未找到分析时间列，直接返回新数据")
                return df_results.copy()
                
        except Exception as e:
            print(f"⚠️ 合并数据时出错: {e}")
            return df_results.copy()
    
    def save_to_excel(self, df_results, additional_info=None):
        """
        保存分析结果到Excel（包含图表、格式化、可视化）
        增量保存：不覆盖前面的数据，只保留每天最后一次分析
        
        参数:
        df_results: 分析结果的DataFrame
        additional_info: 额外信息字典
        """
        if df_results.empty:
            print("❌ 没有数据可保存")
            return False
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            from openpyxl.utils import get_column_letter
            
            # 尝试合并现有数据
            merged_results = self._merge_with_existing_data(df_results)
            
            # 创建Excel写入器
            with pd.ExcelWriter(self.output_file, engine='openpyxl', mode='w') as writer:
                # 1. 保存详细分析结果（合并后的数据）
                merged_results.to_excel(writer, sheet_name='股票分析详情', index=False)
                
                # 2. 创建汇总统计表
                summary_data = self._create_summary_sheet(merged_results)
                summary_data.to_excel(writer, sheet_name='分析汇总', index=False)
                
                # 3. 创建推荐排序表
                if '操作建议' in merged_results.columns:
                    recommendation_data = self._create_recommendation_sheet(merged_results)
                    recommendation_data.to_excel(writer, sheet_name='推荐排序', index=False)
                
                # 4. 创建风险等级分布表
                if '风险等级' in merged_results.columns:
                    risk_dist = self._create_risk_distribution(merged_results)
                    risk_dist.to_excel(writer, sheet_name='风险等级分布', index=False)
                
                # 5. 创建涨跌幅分析表
                if '涨跌幅(%)' in merged_results.columns:
                    change_analysis = self._create_change_analysis(merged_results)
                    change_analysis.to_excel(writer, sheet_name='涨跌幅分析', index=False)
                
                # 6. 添加统计信息
                if additional_info:
                    info_df = pd.DataFrame([additional_info])
                    info_df.to_excel(writer, sheet_name='分析信息', index=False)
                
                # 获取工作簿对象进行格式化
                workbook = writer.book
                
                # 应用格式化
                self._format_detail_sheet(workbook, merged_results)
                self._format_summary_sheet(workbook, summary_data)
                if '操作建议' in merged_results.columns:
                    self._format_recommendation_sheet(workbook)
                if '风险等级' in merged_results.columns:
                    self._add_risk_charts(workbook, merged_results)
                if '涨跌幅(%)' in merged_results.columns:
                    self._add_change_charts(workbook, merged_results)
            
            print(f"✅ 分析结果已保存到: {self.output_file}")
            print(f"📊 包含工作表:")
            print(f"  1. 股票分析详情 - 详细分析结果 (含格式化, 共{len(merged_results)}条)")
            print(f"  2. 分析汇总 - 统计汇总信息")
            print(f"  3. 推荐排序 - 按建议排序")
            print(f"  4. 风险等级分布 - 风险分布及图表")
            print(f"  5. 涨跌幅分析 - 涨跌幅统计及图表")
            if additional_info:
                print(f"  6. 分析信息 - 分析参数信息")
            print(f"💡 特色: 包含条件格式化、图表和详细统计")
            print(f"💾 注: 已保留历史分析数据，每日只保留最后一次分析")
            
            return True
            
        except Exception as e:
            print(f"❌ 保存Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_summary_sheet(self, df_results):
        """创建汇总统计表"""
        if df_results.empty:
            return pd.DataFrame()
        
        summary_data = []
        
        # 基本统计
        total_stocks = len(df_results)
        avg_price = df_results['当前价格'].mean()
        avg_change = df_results['涨跌幅(%)'].mean()
        
        # 涨跌统计
        up_stocks = len(df_results[df_results['涨跌幅(%)'] > 0])
        down_stocks = len(df_results[df_results['涨跌幅(%)'] < 0])
        flat_stocks = len(df_results[df_results['涨跌幅(%)'] == 0])
        
        # 风险等级统计
        if '风险等级' in df_results.columns:
            risk_counts = df_results['风险等级'].value_counts().to_dict()
        else:
            risk_counts = {}
        
        # 操作建议统计
        if '操作建议' in df_results.columns:
            suggestion_counts = df_results['操作建议'].value_counts().to_dict()
        else:
            suggestion_counts = {}
        
        # 添加汇总数据
        summary_data.append({
            '统计项': '分析股票总数',
            '数值': total_stocks,
            '说明': '本次成功分析的股票数量'
        })
        
        summary_data.append({
            '统计项': '平均价格',
            '数值': round(avg_price, 2),
            '说明': '所有股票的平均当前价格'
        })
        
        summary_data.append({
            '统计项': '平均涨跌幅',
            '数值': f"{round(avg_change, 2)}%",
            '说明': '所有股票的平均涨跌幅'
        })
        
        summary_data.append({
            '统计项': '上涨股票数',
            '数值': up_stocks,
            '说明': f"占比: {up_stocks/total_stocks*100:.1f}%"
        })
        
        summary_data.append({
            '统计项': '下跌股票数',
            '数值': down_stocks,
            '说明': f"占比: {down_stocks/total_stocks*100:.1f}%"
        })
        
        summary_data.append({
            '统计项': '平盘股票数',
            '数值': flat_stocks,
            '说明': f"占比: {flat_stocks/total_stocks*100:.1f}%"
        })
        
        # 添加风险等级统计
        for risk_level, count in risk_counts.items():
            summary_data.append({
                '统计项': f'风险等级: {risk_level}',
                '数值': count,
                '说明': f"占比: {count/total_stocks*100:.1f}%"
            })
        
        return pd.DataFrame(summary_data)
    
    def _create_recommendation_sheet(self, df_results):
        """创建推荐排序表"""
        if df_results.empty:
            return pd.DataFrame()
        
        # 复制数据
        recommendation_df = df_results.copy()
        
        # 添加推荐分数（简单逻辑）
        def calculate_score(row):
            score = 0
            
            # 基于涨跌幅
            change = row.get('涨跌幅(%)', 0)
            if change > 5:
                score += 30
            elif change > 0:
                score += 20
            elif change > -5:
                score += 10
            
            # 基于风险等级
            risk = row.get('风险等级', '')
            if risk == '低' or risk == '中低':
                score += 30
            elif risk == '中等':
                score += 20
            elif risk == '中高':
                score += 10
            
            # 基于价格（假设价格适中的股票更好）
            price = row.get('当前价格', 0)
            if 10 <= price <= 100:
                score += 20
            elif 0 < price < 10:
                score += 15
            elif 100 < price < 500:
                score += 10
            
            return score
        
        if '涨跌幅(%)' in recommendation_df.columns:
            recommendation_df['推荐分数'] = recommendation_df.apply(calculate_score, axis=1)
            
            # 按推荐分数排序
            recommendation_df = recommendation_df.sort_values('推荐分数', ascending=False)
            
            # 只保留关键列
            key_columns = ['股票代码', '股票名称', '当前价格', '涨跌幅(%)', 
                          '风险等级', '操作建议', '推荐分数', '分析时间']
            
            # 只选择存在的列
            available_columns = [col for col in key_columns if col in recommendation_df.columns]
            recommendation_df = recommendation_df[available_columns]
        
        return recommendation_df
    
    def _create_risk_distribution(self, df_results):
        """创建风险等级分布表"""
        if '风险等级' not in df_results.columns:
            return pd.DataFrame()
        
        risk_counts = df_results['风险等级'].value_counts().sort_index()
        total = len(df_results)
        
        risk_data = []
        for risk_level, count in risk_counts.items():
            percentage = (count / total) * 100
            risk_data.append({
                '风险等级': risk_level,
                '股票数量': count,
                '占比(%)': round(percentage, 1),
                '占比度': f"{'█' * (count // max(1, total // 10))}{'░' * (10 - count // max(1, total // 10))}"
            })
        
        return pd.DataFrame(risk_data)
    
    def _create_change_analysis(self, df_results):
        """创建涨跌幅分析表"""
        if '涨跌幅(%)' not in df_results.columns:
            return pd.DataFrame()
        
        change_data = df_results['涨跌幅(%)'].copy()
        total = len(change_data)
        
        # 创建分组统计
        bins = [-float('inf'), -10, -5, 0, 5, 10, float('inf')]
        labels = ['极度下跌(≤-10%)', '大幅下跌(-5%~-10%)', '小幅下跌(-5%~0%)', 
                  '小幅上涨(0%~5%)', '大幅上涨(5%~10%)', '极度上涨(>10%)']
        
        change_bins = pd.cut(change_data, bins=bins, labels=labels, right=False)
        bin_counts = change_bins.value_counts().sort_index()
        
        analysis_data = []
        for label, count in bin_counts.items():
            percentage = (count / total) * 100
            analysis_data.append({
                '涨跌幅区间': label,
                '股票数量': count,
                '占比(%)': round(percentage, 1)
            })
        
        return pd.DataFrame(analysis_data)
    
    def _format_detail_sheet(self, workbook, df_results):
        """对详情表进行格式化（条件格式化、颜色等）"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.utils import get_column_letter
            
            ws = workbook['股票分析详情']
            
            # 定义样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 格式化表头
            for col_num, col_title in enumerate(df_results.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # 设置列宽
                if col_title in ['股票名称', '分析结果']:
                    ws.column_dimensions[get_column_letter(col_num)].width = 25
                else:
                    ws.column_dimensions[get_column_letter(col_num)].width = 15
            
            # 对数据行应用格式
            for row_num in range(2, len(df_results) + 2):
                for col_num, col_title in enumerate(df_results.columns, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 条件格式化：涨跌幅
                    if col_title == '涨跌幅(%)':
                        try:
                            value = float(cell.value) if cell.value else 0
                            if value > 5:
                                cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                                cell.font = Font(bold=True, color='FF0000')
                            elif value > 0:
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            elif value < -5:
                                cell.fill = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')
                                cell.font = Font(bold=True, color='FFFFFF')
                            elif value < 0:
                                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        except:
                            pass
                    
                    # 条件格式化：风险等级
                    elif col_title == '风险等级':
                        risk_colors = {
                            '低': 'C6EFCE',
                            '中低': 'B7DEE8',
                            '中': 'FFE699',
                            '中高': 'F4B084',
                            '高': 'F8696B',
                            '极高': '800000'
                        }
                        risk = str(cell.value).strip() if cell.value else ''
                        if risk in risk_colors:
                            color = risk_colors[risk]
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            if risk in ['高', '极高']:
                                cell.font = Font(bold=True, color='FFFFFF')
        except Exception as e:
            print(f"⚠️ 格式化详情表失败: {e}")
    
    def _format_summary_sheet(self, workbook, summary_data):
        """对汇总表进行格式化"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            ws = workbook['分析汇总']
            
            header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 格式化表头
            for col_num in range(1, len(summary_data.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 格式化数据行
            for row_num in range(2, len(summary_data) + 2):
                for col_num in range(1, len(summary_data.columns) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    if col_num == 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 交替行颜色
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 设置列宽
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 30
        except Exception as e:
            print(f"⚠️ 格式化汇总表失败: {e}")
    
    def _format_recommendation_sheet(self, workbook):
        """对推荐表进行格式化"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            ws = workbook['推荐排序']
            
            header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 格式化表头
            for col_num in range(1, 9):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 获取数据行数
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 推荐分数列色彩化
                    if cell.column == 7:  # 推荐分数列
                        try:
                            score = float(cell.value) if cell.value else 0
                            if score >= 80:
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                cell.font = Font(bold=True, color='006100')
                            elif score >= 60:
                                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        except:
                            pass
        except Exception as e:
            print(f"⚠️ 格式化推荐表失败: {e}")
    
    def _add_risk_charts(self, workbook, df_results):
        """添加风险等级分布图表"""
        try:
            from openpyxl.chart import PieChart, BarChart, Reference
            
            if '风险等级' not in df_results.columns:
                return
            
            ws = workbook['风险等级分布']
            
            # 创建饼图
            pie = PieChart()
            pie.title = "风险等级分布（饼图）"
            pie.style = 10
            
            # 数据范围（假设数据在A2:B*)
            labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            
            ws.add_chart(pie, "D2")
            
            # 创建柱状图
            bar = BarChart()
            bar.title = "风险等级分布（柱状图）"
            bar.x_axis.title = "风险等级"
            bar.y_axis.title = "股票数量"
            bar.style = 10
            
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            
            ws.add_chart(bar, "D18")
        except Exception as e:
            print(f"⚠️ 添加风险图表失败: {e}")
    
    def _add_change_charts(self, workbook, df_results):
        """添加涨跌幅分析图表"""
        try:
            from openpyxl.chart import BarChart, LineChart, Reference
            
            if '涨跌幅(%)' not in df_results.columns:
                return
            
            ws = workbook['涨跌幅分析']
            
            # 创建柱状图
            bar = BarChart()
            bar.title = "涨跌幅区间分布"
            bar.x_axis.title = "涨跌幅区间"
            bar.y_axis.title = "股票数量"
            bar.style = 10
            
            # 数据范围
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            
            ws.add_chart(bar, "D2")
            
            # 创建二级轴处理（占比）
            line = LineChart()
            line.title = "涨跌幅占比分析"
            
            data2 = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            line.add_data(data2, titles_from_data=True)
            line.set_categories(cats)
            line.y_axis.title = "占比 (%)"
            
            ws.add_chart(line, "D18")
        except Exception as e:
            print(f"⚠️ 添加涨跌幅图表失败: {e}")
    
    def run_batch_analysis(self, stock_list, output_file=None):
        """
        运行批量分析的主函数
        
        参数:
        stock_list: 股票代码列表
        output_file: 输出文件名（可选）
        """
        if output_file:
            self.output_file = output_file
        
        print("=" * 60)
        print("📈 批量股票分析系统")
        print("=" * 60)
        
        # 记录开始时间
        start_time = datetime.now()
        
        # 执行批量分析
        df_results = self.analyze_stocks_batch(stock_list)
        
        # 记录结束时间
        end_time = datetime.now()
        time_cost = (end_time - start_time).total_seconds()
        
        # 准备额外信息
        additional_info = {
            '分析开始时间': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            '分析结束时间': end_time.strftime('%Y-%m-%d %H:%M:%S'),
            '分析耗时(秒)': round(time_cost, 2),
            '分析股票数量': len(stock_list),
            '成功分析数量': len(df_results) if not df_results.empty else 0,
            '输出文件': self.output_file
        }
        
        # 保存到Excel
        if not df_results.empty:
            self.save_to_excel(df_results, additional_info)
            
            # 显示简要结果
            print(f"\n🎯 分析摘要:")
            print(f"  耗时: {time_cost:.1f}秒")
            print(f"  平均每只股票: {time_cost/len(stock_list):.1f}秒")
            
            if '涨跌幅(%)' in df_results.columns:
                top_gainers = df_results.nlargest(10, '涨跌幅(%)')
                print(f"\n📈 涨幅前十:")
                for idx, row in top_gainers.iterrows():
                    print(f"  {row['股票代码']} {row['股票名称']}: {row['涨跌幅(%)']}%")
            
            if '推荐分数' in df_results.columns:
                top_recommended = df_results.nlargest(3, '推荐分数')
                print(f"\n⭐ 推荐前三:")
                for idx, row in top_recommended.iterrows():
                    print(f"  {row['股票代码']} {row['股票名称']}: {row['操作建议']}")
        
        return df_results

# ========== 使用示例 ==========
def main():
    """使用示例"""
    
    # 示例1：直接定义股票列表
    example_stocks = [
        '600312',   # 平高电气
        '600635',   # 大众公用
        '002131',   # 利欧股份
        '002415',   # 海康威视   
        '600519',   # 贵州茅台
        '600436',   # 片仔癀
        '600309'    # 万华化学
    ]
    
    # 示例2：从文件读取股票列表（可选）
    def read_stocks_from_file(filename):
        """从文本文件读取股票列表"""
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                stocks = [line.strip() for line in f if line.strip()]
            return stocks
        except:
            return []
    
    # 初始化分析器
    analyzer = BatchStockAnalyzer(output_file='批量股票分析报告.xlsx')
    
    # 选择股票列表来源
    print("请选择股票列表来源:")
    print("1. 使用示例股票列表")
    print("2. 从文件读取股票列表")
    print("3. 手动输入股票列表")
    
    # choice = input("请输入选择 (1/2/3): ").strip()
    
    stock_list = example_stocks
    print(f"使用示例股票列表，共 {len(stock_list)} 只股票")
    
    # if choice == '1':
    #     stock_list = example_stocks
    #     print(f"使用示例股票列表，共 {len(stock_list)} 只股票")
    # elif choice == '2':
    #     filename = input("请输入股票列表文件路径: ").strip()
    #     stock_list = read_stocks_from_file(filename)
    #     if stock_list:
    #         print(f"从文件读取 {len(stock_list)} 只股票")
    #     else:
    #         print("文件读取失败，使用示例股票列表")
    #         stock_list = example_stocks
    # else:
    #     # 手动输入
    #     input_str = input("请输入股票代码（用逗号或空格分隔）: ").strip()
    #     if ',' in input_str:
    #         stock_list = [code.strip() for code in input_str.split(',') if code.strip()]
    #     else:
    #         stock_list = [code.strip() for code in input_str.split() if code.strip()]
    #     print(f"手动输入 {len(stock_list)} 只股票")
    
    if not stock_list:
        print("❌ 股票列表为空，程序退出")
        return
    
    # 确认股票列表
    print(f"\n📋 将要分析的股票 ({len(stock_list)} ):")
    for i, code in enumerate(stock_list[:20]):  # 只显示前20只
        print(f"  {i+1:2d}. {code}", end='\n' if (i+1) % 5 == 0 else '  ')
    
    if len(stock_list) > 20:
        print(f"  ... 等 {len(stock_list)} 只股票")
    
    # 确认开始分析
    # confirm = input("\n是否开始分析? (y/n): ").strip().lower()
    # if confirm != 'y':
    #     print("分析已取消")
    #     return
    
    # 运行批量分析
    df_results = analyzer.run_batch_analysis(stock_list)
    
    print(f"\n✅ 批量分析完成!")
    print(f"请查看输出文件: {analyzer.output_file}")

if __name__ == "__main__":
    # 检查必要库
    try:
        import pandas as pd
        import requests
        print("✓ 必要库已安装")
    except ImportError as e:
        print(f"❌ 缺少必要库: {e}")
        print("请运行: pip install pandas requests openpyxl")
        exit(1)
    
    # 运行主程序
    main()