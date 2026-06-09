#!/usr/bin/env python3
"""Check i18n image differences between zh_TW and mainland client trees."""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import hashlib
import html
import http.server
import json
import shutil
import os
import re
import socketserver
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Callable, Sequence
from xml.etree import ElementTree as ET

BEIJING_TZ = dt.timezone(dt.timedelta(hours=8))
SCRIPT_VERSION = "1.0.2"
SCRIPT_UPDATED_AT = "2026-06-05 09:21:59 +08:00"

IMAGE_EXTENSIONS = {".dds", ".tga"}
STATE_FILE = ".check_i18n_images_state"
OCR_CACHE_FILE = ".ocr_cache.json"
OCR_OPERATION_IGNORE = "ignore"
DEFAULT_THUMBNAIL_ISSUES = {
    "__has_detail__",
}
_RAPIDOCR_ENGINE = None
_RAPIDOCR_LOCAL = threading.local()
TEXT_PROJECT_NAME = "\u5251\u7f51\u4e09"
TEXT_REPORT_TITLE = "\u591a\u8bed\u8a00\u56fe\u7247\u68c0\u67e5\u6c47\u603b"
TEXT_RUN_RESULT = "\u672c\u8f6e\u6267\u884c\u7ed3\u679c"
TEXT_MISSING_ISSUE = "\u7591\u4f3c\u5e9f\u9664\u6587\u4ef6"
TEXT_CHANGED_ISSUE = "\u4fee\u6539\u65f6\u95f4\u5f02\u5e38"
TEXT_NEW_TEXT_ISSUE = "\u65b0\u589e\u5e26\u6587\u5b57\u56fe\u7247"
TEXT_NEW_NO_TEXT_ISSUE = "\u65b0\u589e\u65e0\u6587\u5b57\u56fe\u7247"
TEXT_REPORT_DETAIL = "\u62a5\u544a\u8be6\u60c5"
TEXT_OTHER_ISSUE = "\u5176\u4ed6\u95ee\u9898"
TEXT_CATEGORY = "\u7c7b\u522b"
TEXT_ISSUE_TYPE = "\u95ee\u9898\u7c7b\u578b"
TEXT_RELATIVE_PATH = "\u76f8\u5bf9\u8def\u5f84"
TEXT_I18N_IMAGE = "\u56fd\u9645\u7248\u56fe\u7247"
TEXT_MAINLAND_IMAGE = "\u9646\u7248\u56fe\u7247"
TEXT_I18N_MODIFIED = "\u56fd\u9645\u7248\u4fee\u6539\u65f6\u95f4"
TEXT_MAINLAND_MODIFIED = "\u9646\u7248\u4fee\u6539\u65f6\u95f4"
TEXT_DETAIL = "\u8bf4\u660e"
TEXT_OPERATION = "\u64cd\u4f5c"
TEXT_IGNORE = "\u5ffd\u7565"
TEXT_IGNORED = "\u5df2\u5ffd\u7565"
TEXT_I18N = "\u56fd\u9645\u7248"
TEXT_MAINLAND = "\u9646\u7248"
TEXT_NONE = "\u65e0"
TEXT_ITEM = "\u9879"
TEXT_OCR_TEXT = "\u8bc6\u522b\u6587\u5b57"
_RUN_LOG_FILE: Path | None = None
_OCR_CANDIDATE_FILE: Path | None = None
_OCR_CANDIDATES: list[ImageFile] = []

@dataclass(frozen=True)
class ImageFile:
    relative_path: str
    full_path: str
    modified_at: dt.datetime
    category: str
    created_at: dt.datetime | None = None


@dataclass(frozen=True)
class Finding:
    category: str
    issue: str
    relative_path: str
    i18n_path: str
    mainland_path: str
    i18n_modified_at: dt.datetime | None
    mainland_modified_at: dt.datetime | None
    detail: str
    mainland_created_at: dt.datetime | None = None
    mainland_ocr_text: str | None = None
    i18n_ocr_text: str | None = None
    # ok / error / missing / no_cn_text
    translation_status: str | None = None
    translation_note: str | None = None
    operation: str | None = None
    mainland_md5: str | None = None


def normalize_rel(path: str) -> str:
    return str(PurePosixPath(path.replace("\\", "/"))).lstrip("/")


def compare_key(relative_path: str) -> str:
    return normalize_rel(relative_path).casefold()


def image_match_key(relative_path: str) -> str:
    normalized = normalize_rel(relative_path)
    path = PurePosixPath(normalized)
    if path.suffix.lower() in IMAGE_EXTENSIONS:
        return compare_key(str(path.with_suffix("")))
    return compare_key(normalized)


def is_image(path: str) -> bool:
    return PurePosixPath(path).suffix.lower() in IMAGE_EXTENSIONS


def to_aware_utc(value: dt.datetime) -> dt.datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=dt.timezone.utc)
    return value.astimezone(dt.timezone.utc)


def _init_run_outputs(now: dt.datetime | None = None) -> tuple[Path, Path]:
    global _RUN_LOG_FILE, _OCR_CANDIDATE_FILE, _OCR_CANDIDATES
    timestamp = (now or dt.datetime.now()).strftime("%Y%m%d_%H%M%S")
    log_dir = Path("Log")
    log_dir.mkdir(parents=True, exist_ok=True)
    _RUN_LOG_FILE = log_dir / f"{timestamp}.log"
    _OCR_CANDIDATE_FILE = log_dir / f"{timestamp}_ocr_images.txt"
    _OCR_CANDIDATES = []
    _RUN_LOG_FILE.write_text("", encoding="utf-8")
    _OCR_CANDIDATE_FILE.write_text("", encoding="utf-8")
    return _RUN_LOG_FILE, _OCR_CANDIDATE_FILE


def log_step(message: str) -> None:
    timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line, file=sys.stderr)
    if _RUN_LOG_FILE is not None:
        try:
            with _RUN_LOG_FILE.open("a", encoding="utf-8") as f:
                f.write(line + "\n")
        except OSError:
            pass


def format_version_line(now: dt.datetime | None = None) -> str:
    current = now or dt.datetime.now(BEIJING_TZ)
    if current.tzinfo is None:
        current = current.replace(tzinfo=BEIJING_TZ)
    return f"version = {SCRIPT_VERSION}，time = {current.astimezone(BEIJING_TZ).strftime('%Y-%m-%d %H:%M:%S %z')[:-2]}:00"


def _record_ocr_candidate(image: ImageFile) -> None:
    _OCR_CANDIDATES.append(image)


def _write_ocr_candidate_list() -> None:
    if _OCR_CANDIDATE_FILE is None:
        return
    lines = [
        f"{index}\t{image.category}\t{image.relative_path}\t{image.full_path}"
        for index, image in enumerate(_OCR_CANDIDATES, 1)
    ]
    _OCR_CANDIDATE_FILE.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def log_new_no_text_findings(findings: Sequence[Finding]) -> None:
    items = [f for f in findings if f.issue == "mainland_new_no_text"]
    log_step(f"新增无文字图片列表: count={len(items)}")
    for index, finding in enumerate(items, 1):
        log_step(f"新增无文字图片 {index}: {finding.relative_path} | {finding.mainland_path}")


def default_ocr_workers() -> int:
    return 1


def compare_category(
    category: str,
    i18n_files: dict[str, ImageFile],
    mainland_files: dict[str, ImageFile],
    last_check_at: dt.datetime | None,
    text_detector: Callable[[ImageFile], bool],
    ocr_workers: int = 1,
) -> tuple[list[Finding], dict[str, int]]:
    findings: list[Finding] = []
    normal_synced = 0
    new_no_text = 0

    new_mainland_files: list[tuple[str, ImageFile]] = []
    for rel, mainland_file in sorted(mainland_files.items()):
        i18n_file = i18n_files.get(rel)
        if i18n_file is None:
            new_mainland_files.append((rel, mainland_file))
            continue
        if to_aware_utc(i18n_file.modified_at) < to_aware_utc(mainland_file.modified_at):
            findings.append(
                Finding(
                    category,
                    "mainland_changed",
                    i18n_file.relative_path,
                    i18n_file.full_path,
                    mainland_file.full_path,
                    i18n_file.modified_at,
                    mainland_file.modified_at,
                    "陆版文件修改时间晚于国际版",
                )
            )
        else:
            normal_synced += 1

    for rel, i18n_file in sorted(i18n_files.items()):
        if rel in mainland_files:
            continue
        findings.append(
            Finding(
                category,
                "mainland_missing",
                i18n_file.relative_path,
                i18n_file.full_path,
                "",
                i18n_file.modified_at,
                None,
                "陆版无对应文件，国际版文件可能已废弃",
            )
        )

    new_total = len(new_mainland_files)
    log_step(f"OCR 候选图片: category={category or '-'} count={new_total}")
    for _rel, mainland_file in new_mainland_files:
        _record_ocr_candidate(mainland_file)
    progress_started_at = time.monotonic()
    progress_line_length = 0

    def detect_text(index_and_image: tuple[int, ImageFile]) -> tuple[int, ImageFile, bool, str | None, str, str | None, str | None]:
        index, image = index_and_image
        has_text = text_detector(image)
        ocr_text_getter = getattr(text_detector, "ocr_text_for", None)
        ocr_text = ocr_text_getter(image) if callable(ocr_text_getter) else None
        operation_getter = getattr(text_detector, "operation_for", None)
        operation = operation_getter(image) if callable(operation_getter) else None
        md5_getter = getattr(text_detector, "md5_for", None)
        file_md5 = md5_getter(image) if callable(md5_getter) else None
        return index, image, has_text, ocr_text, threading.current_thread().name, operation, file_md5

    def print_ocr_progress(done_count: int, image: ImageFile, worker_name: str) -> None:
        nonlocal progress_line_length
        elapsed_minutes = int((time.monotonic() - progress_started_at) // 60)
        elapsed_hours = elapsed_minutes // 60
        elapsed_remainder_minutes = elapsed_minutes % 60
        progress_line = format_progress_line(
            done_count,
            new_total,
            elapsed_hours,
            elapsed_remainder_minutes,
            image.full_path,
            worker_name,
        )
        padding = max(0, progress_line_length - len(progress_line))
        progress_line_length = len(progress_line)
        print(f"\r{progress_line}{' ' * padding}", end="", file=sys.stderr, flush=True)

    worker_count = max(1, int(ocr_workers or 1))
    indexed_images = [
        (index, mainland_file)
        for index, (_rel, mainland_file) in enumerate(new_mainland_files, 1)
    ]
    if worker_count == 1 or len(indexed_images) <= 1:
        detected_results = []
        for item in indexed_images:
            result = detect_text(item)
            detected_results.append(result)
            _index, image, _has_text, _ocr_text, worker_name, _operation, _file_md5 = result
            print_ocr_progress(len(detected_results), image, worker_name)
    else:
        log_step(f"启动 OCR 并发: workers={worker_count} count={new_total}")
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = [executor.submit(detect_text, item) for item in indexed_images]
            detected_results = []
            for future in as_completed(futures):
                result = future.result()
                detected_results.append(result)
                _index, image, _has_text, _ocr_text, worker_name, _operation, _file_md5 = result
                print_ocr_progress(len(detected_results), image, worker_name)
        detected_results.sort(key=lambda item: item[0])

    for _progress_index, mainland_file, has_text, mainland_ocr_text, _worker_name, operation, file_md5 in detected_results:
        if has_text:
            findings.append(
                Finding(
                    category,
                    "mainland_new_with_text",
                    mainland_file.relative_path,
                    "",
                    mainland_file.full_path,
                    None,
                    mainland_file.modified_at,
                    "陆版新增图片且检测到文字",
                    mainland_created_at=mainland_file.created_at,
                    mainland_ocr_text=mainland_ocr_text,
                    operation=operation,
                    mainland_md5=file_md5,
                )
            )
        else:
            new_no_text += 1
            findings.append(
                Finding(
                    category,
                    "mainland_new_no_text",
                    mainland_file.relative_path,
                    "",
                    mainland_file.full_path,
                    None,
                    mainland_file.modified_at,
                    "陆版新增图片，OCR 未检测到文字，建议人工复核",
                    mainland_created_at=mainland_file.created_at,
                    operation=operation,
                    mainland_md5=file_md5,
                )
            )
    if new_total:
        print(file=sys.stderr)
    return findings, {"normal_synced": normal_synced, "new_no_text": new_no_text}


def format_progress_line(
    progress_index: int,
    total: int,
    elapsed_hours: int,
    elapsed_minutes: int,
    full_path: str,
    worker_name: str | None = None,
) -> str:
    worker = f"OCR线程：{worker_name}；" if worker_name else ""
    return (
        f"当前分析进度：{progress_index}/{total}，"
        f"共执行：{elapsed_hours:02d}小时{elapsed_minutes:02d}分钟；"
        f"{worker}"
        f"路径：{full_path}"
    )


def apply_max_file_sample(
    i18n_files: dict[str, ImageFile],
    mainland_files: dict[str, ImageFile],
    max_files: int | None,
) -> tuple[dict[str, ImageFile], dict[str, ImageFile]]:
    if max_files is None or max_files < 0:
        return i18n_files, mainland_files
    sampled_keys = set(sorted(set(i18n_files) | set(mainland_files))[:max_files])
    return (
        {key: value for key, value in sorted(i18n_files.items()) if key in sampled_keys},
        {key: value for key, value in sorted(mainland_files.items()) if key in sampled_keys},
    )


def _whitelist_ignored_prefixes(pair: dict[str, object]) -> list[str]:
    raw_whitelist = pair.get("whitelist", [])
    whitelist: list[str] = []
    for item in raw_whitelist:
        if isinstance(item, dict):
            continue
        parsed_rule = _parse_inline_whitelist_rule(str(item))
        if parsed_rule is None:
            whitelist.append(normalize_rel(str(item)))

    relative_roots = [
        normalize_rel(str(pair.get("mainland_relative", ""))),
        normalize_rel(str(pair.get("i18n_relative", ""))),
    ]
    ignored_prefixes: list[str] = []
    for item in whitelist:
        for root in relative_roots:
            if root and compare_key(item).startswith(compare_key(root) + "/"):
                ignored_prefixes.append(compare_key(item[len(root):].lstrip("/")))
                break
        else:
            ignored_prefixes.append(compare_key(item))
    return ignored_prefixes


def _is_whitelisted_path(relative_path: str, pair: dict[str, object]) -> bool:
    path_key = compare_key(relative_path)
    if not path_key:
        return False
    if any(path_key == prefix or path_key.startswith(prefix + "/") for prefix in _whitelist_ignored_prefixes(pair)):
        return True

    relative_roots = [
        normalize_rel(str(pair.get("mainland_relative", ""))),
        normalize_rel(str(pair.get("i18n_relative", ""))),
    ]
    for item in pair.get("whitelist", []):
        if isinstance(item, dict):
            rule = item
        else:
            rule = _parse_inline_whitelist_rule(str(item))
            if rule is None:
                continue
        if _matches_whitelist_rule(relative_path, rule, relative_roots):
            return True
    return False


def scan_local(directory: str, pair: dict[str, object] | None = None) -> dict[str, ImageFile]:
    base_path = Path(directory)
    result: dict[str, ImageFile] = {}
    if not base_path.exists():
        return result

    scanned = 0
    skipped = 0
    for root, dirs, files in os.walk(base_path):
        root_path = Path(root)
        if pair is not None:
            kept_dirs = []
            for directory_name in dirs:
                rel_dir = normalize_rel((root_path / directory_name).relative_to(base_path).as_posix())
                if not _is_whitelisted_path(rel_dir, pair):
                    kept_dirs.append(directory_name)
            dirs[:] = kept_dirs
        for file_name in files:
            file = root_path / file_name
            rel = normalize_rel(file.relative_to(base_path).as_posix())
            if pair is not None and _is_whitelisted_path(rel, pair):
                continue
            try:
                if not file.is_file():
                    continue
            except OSError as e:
                skipped += 1
                if skipped <= 10:
                    print(f"WARN: 跳过无法访问的文件: {file}: {e}", file=sys.stderr)
                continue
            if not is_image(rel):
                continue
            try:
                stat = file.stat()
            except OSError as e:
                skipped += 1
                if skipped <= 10:
                    print(f"WARN: 跳过无法获取状态的文件: {file}: {e}", file=sys.stderr)
                continue
            modified_at = dt.datetime.fromtimestamp(stat.st_mtime, dt.timezone.utc)
            created_at = dt.datetime.fromtimestamp(stat.st_ctime, dt.timezone.utc)
            result.setdefault(
                image_match_key(rel),
                ImageFile(rel, str(file), modified_at, "", created_at),
            )
            scanned += 1
            if scanned % 5000 == 0:
                print(f"  已扫描 {scanned} 张图片...", file=sys.stderr)
    if skipped > 10:
        print(f"  共跳过 {skipped} 个无法访问的文件", file=sys.stderr)
    return result


def run_svn(args: Sequence[str]) -> bytes:
    proc = subprocess.run(
        ["svn", *args],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    if proc.returncode != 0:
        stderr = proc.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(stderr or f"svn {' '.join(args)} failed")
    return proc.stdout


def scan_svn(url: str) -> dict[str, ImageFile]:
    try:
        xml = run_svn(["list", "--xml", "-R", url])
    except RuntimeError as exc:
        print(f"WARN: 跳过无法读取的 SVN 路径: {url}: {exc}", file=sys.stderr)
        return {}

    doc = ET.fromstring(xml)
    result: dict[str, ImageFile] = {}
    for entry in doc.findall(".//entry"):
        if entry.get("kind") != "file":
            continue
        rel = normalize_rel(entry.findtext("name") or "")
        if not is_image(rel):
            continue
        date_text = entry.findtext("commit/date")
        if not date_text:
            continue
        modified_at = dt.datetime.fromisoformat(date_text.replace("Z", "+00:00"))
        result.setdefault(
            image_match_key(rel),
            ImageFile(rel, url.rstrip("/") + "/" + rel.strip("/"), modified_at, ""),
        )
    return result


def default_text_detector(image: ImageFile) -> bool:
    return False


def assume_text_detector(image: ImageFile) -> bool:
    return True


def _prepare_ocr_sources(source: str) -> tuple[list[str], list[str]]:
    suffix = PurePosixPath(source).suffix.lower()
    if suffix not in {".tga", ".dds"}:
        return [source], []
    try:
        from PIL import Image, ImageEnhance, ImageOps
    except ImportError:
        return [source], []

    sources = [source]
    temps: list[str] = []

    def add_temp(img) -> None:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.close()
        img.save(tmp.name)
        sources.append(tmp.name)
        temps.append(tmp.name)

    try:
        with Image.open(source) as img:
            rgb = img.convert("RGB")
            mask = rgb.convert("L").point(lambda v: 255 if v > 20 else 0)
            bbox = mask.getbbox()
            cropped = rgb.crop(bbox) if bbox else rgb
            add_temp(cropped.convert("RGBA"))
            is_vertical_asset = cropped.height >= cropped.width * 2
            if is_vertical_asset:
                rotated = cropped.rotate(90, expand=True)
                add_temp(rotated.convert("RGBA"))
            up3 = cropped.resize((cropped.width * 3, cropped.height * 3))
            add_temp(up3.convert("RGBA"))
            if is_vertical_asset:
                add_temp(up3.rotate(90, expand=True).convert("RGBA"))

            gray = ImageOps.grayscale(up3)
            gray = ImageEnhance.Contrast(gray).enhance(2.5)
            add_temp(gray.point(lambda v: 0 if v > 35 else 255))
    except Exception:
        return sources, temps
    return sources, temps


def _ocr_text_score(text: str) -> int:
    chinese = len(re.findall(r"[一-鿿]", text))
    ascii_words = len(re.findall(r"[A-Za-z0-9]", text))
    noise = len(re.findall(r"[�Î˝ˋ′`\\/_{}\[\]<>]", text))
    return chinese * 10 + ascii_words - noise * 2


def _normalize_ocr_text(text: str) -> str:
    replacements = {
        "加入巧帮": "加入丐帮",
    }
    for wrong, right in replacements.items():
        text = text.replace(wrong, right)
    return text


def _chapter_filename_text_hint(image: ImageFile) -> str | None:
    rel = normalize_rel(image.relative_path or image.full_path)
    if "image/chapterstga/" not in compare_key(rel):
        return None
    stem = PurePosixPath(rel).stem
    if stem.startswith("箭头"):
        return None
    stem = re.sub(r"_HD$", "", stem, flags=re.IGNORECASE)
    stem = stem.replace("_", " ")
    parts = [part for part in re.split(r"[\s·]+", stem) if re.search(r"[\u4e00-\u9fff]", part)]
    return "\n".join(parts) if parts else None


def _merge_ocr_with_filename_hint(text: str | None, image: ImageFile) -> str | None:
    hint = _chapter_filename_text_hint(image)
    if not hint:
        return text
    if not text:
        return hint
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for hint_line in [line.strip() for line in hint.splitlines() if line.strip()]:
        if not any(hint_line in line or line in hint_line for line in lines):
            lines.insert(0, hint_line)
    return "\n".join(lines)


def _rapidocr_engine():
    global _RAPIDOCR_ENGINE
    if threading.current_thread() is threading.main_thread():
        if _RAPIDOCR_ENGINE is None:
            from rapidocr_onnxruntime import RapidOCR

            _RAPIDOCR_ENGINE = RapidOCR()
        return _RAPIDOCR_ENGINE

    engine = getattr(_RAPIDOCR_LOCAL, "engine", None)
    if engine is None:
        from rapidocr_onnxruntime import RapidOCR

        engine = RapidOCR()
        _RAPIDOCR_LOCAL.engine = engine
    return engine


def run_rapidocr(image: ImageFile) -> str | None:
    """Run RapidOCR on an image; return extracted text or None when unavailable or empty."""
    source = image.full_path
    local_temp: str | None = None
    converted_temps: list[str] = []
    try:
        try:
            ocr = _rapidocr_engine()
        except ImportError:
            return None

        if source.lower().startswith(("http://", "https://")):
            suffix = PurePosixPath(source).suffix
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as fp:
                fp.write(run_svn(["cat", source]))
                local_temp = fp.name
                source = fp.name

        sources, converted_temps = _prepare_ocr_sources(source)
        best_text = ""
        original_has_text = False
        for index, candidate in enumerate(sources):
            result, _elapsed = ocr(candidate)
            accepted_texts: list[str] = []
            max_confidence = 0.0
            for item in result or []:
                if len(item) <= 1:
                    continue
                text = str(item[1]).strip()
                if not text:
                    continue
                confidence = 1.0
                if len(item) > 2:
                    try:
                        confidence = float(item[2])
                    except (TypeError, ValueError):
                        confidence = 0.0
                if index == 0 and confidence > 0:
                    original_has_text = True
                max_confidence = max(max_confidence, confidence)
                min_confidence = 0.70 if original_has_text else 0.75
                if index > 0 and confidence < min_confidence:
                    continue
                accepted_texts.append(text)
            text = "\n".join(accepted_texts)
            if text and _ocr_text_score(text) > _ocr_text_score(best_text):
                best_text = text
            if index == 0 and max_confidence >= 0.78 and len(re.findall(r"[一-鿿]", text)) >= 4:
                return _normalize_ocr_text(text)
        return _normalize_ocr_text(best_text) if best_text else None
    except Exception as exc:
        print(f"WARN: RapidOCR 失败: {image.full_path}: {exc}", file=sys.stderr)
        return None
    finally:
        if local_temp:
            try:
                os.unlink(local_temp)
            except OSError:
                pass
        for converted_temp in converted_temps:
            try:
                os.unlink(converted_temp)
            except OSError:
                pass


def run_ocr(image: ImageFile) -> str | None:
    """Run OCR on an image using RapidOCR."""
    return _merge_ocr_with_filename_hint(run_rapidocr(image), image)


def run_ocr_from_path(path: str, category: str = "") -> str | None:
    """Run OCR on a local path or SVN URL; return extracted text."""
    if not path:
        return None
    dummy = ImageFile(
        relative_path="",
        full_path=path,
        modified_at=dt.datetime.now(dt.timezone.utc),
        category=category,
    )
    return run_ocr(dummy)


def _file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_ocr_cache(cache_path: Path) -> dict[str, dict[str, object]]:
    if not cache_path.exists():
        return {}
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_ocr_cache(cache_path: Path, cache: dict[str, dict[str, object]]) -> None:
    cache_path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def update_ocr_cache_operation(
    cache_path: Path,
    relative_path: str,
    file_md5: str,
    operation: str,
) -> None:
    cache = load_ocr_cache(cache_path)
    entry = cache.get(relative_path)
    if not isinstance(entry, dict):
        entry = {}
    entry["md5"] = file_md5
    entry["operation"] = operation
    cache[relative_path] = entry
    save_ocr_cache(cache_path, cache)


def init_ocr_cache_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ocr_cache (
                relative_path TEXT PRIMARY KEY,
                md5 TEXT NOT NULL,
                has_text INTEGER,
                test_str TEXT,
                operation TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS review_operation (
                relative_path TEXT PRIMARY KEY,
                md5 TEXT NOT NULL,
                operation TEXT NOT NULL,
                operator TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def _now_text() -> str:
    return dt.datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S %z")


def migrate_ocr_json_to_sqlite(json_path: Path, db_path: Path) -> int:
    data = load_ocr_cache(json_path)
    init_ocr_cache_db(db_path)
    count = 0
    now_text = _now_text()
    conn = sqlite3.connect(db_path)
    try:
        for relative_path, entry in data.items():
            if not isinstance(entry, dict):
                continue
            file_md5 = entry.get("md5")
            if not isinstance(file_md5, str) or not file_md5:
                continue
            has_text = entry.get("has_text")
            has_text_value = None if has_text is None else int(bool(has_text))
            test_str = entry.get("test_str")
            operation = entry.get("operation")
            conn.execute(
                """
                INSERT INTO ocr_cache(relative_path, md5, has_text, test_str, operation, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(relative_path) DO UPDATE SET
                    md5=excluded.md5,
                    has_text=excluded.has_text,
                    test_str=excluded.test_str,
                    operation=excluded.operation,
                    updated_at=excluded.updated_at
                """,
                (
                    relative_path,
                    file_md5,
                    has_text_value,
                    test_str if isinstance(test_str, str) else "",
                    operation if isinstance(operation, str) else None,
                    now_text,
                ),
            )
            if isinstance(operation, str) and operation:
                conn.execute(
                    """
                    INSERT INTO review_operation(relative_path, md5, operation, operator, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(relative_path) DO UPDATE SET
                        md5=excluded.md5,
                        operation=excluded.operation,
                        operator=excluded.operator,
                        updated_at=excluded.updated_at
                    """,
                    (relative_path, file_md5, operation, "", now_text),
                )
            count += 1
        conn.commit()
    finally:
        conn.close()
    return count


def update_ocr_cache_operation_sqlite(
    db_path: Path,
    relative_path: str,
    file_md5: str,
    operation: str,
    operator: str = "",
) -> None:
    init_ocr_cache_db(db_path)
    now_text = _now_text()
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            INSERT INTO ocr_cache(relative_path, md5, operation, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(relative_path) DO UPDATE SET
                md5=excluded.md5,
                operation=excluded.operation,
                updated_at=excluded.updated_at
            """,
            (relative_path, file_md5, operation, now_text),
        )
        conn.execute(
            """
            INSERT INTO review_operation(relative_path, md5, operation, operator, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(relative_path) DO UPDATE SET
                md5=excluded.md5,
                operation=excluded.operation,
                operator=excluded.operator,
                updated_at=excluded.updated_at
            """,
            (relative_path, file_md5, operation, operator, now_text),
        )
        conn.commit()
    finally:
        conn.close()


def sqlite_ocr_text_detector_factory(
    db_path: Path,
) -> Callable[[ImageFile], bool]:
    init_ocr_cache_db(db_path)
    db_lock = threading.Lock()

    def _entry_for(image: ImageFile, file_md5: str) -> dict[str, object] | None:
        with db_lock:
            conn = sqlite3.connect(db_path)
            try:
                row = conn.execute(
                    """
                    SELECT c.has_text, c.test_str, COALESCE(r.operation, c.operation) AS operation
                    FROM ocr_cache c
                    LEFT JOIN review_operation r
                      ON r.relative_path = c.relative_path AND r.md5 = c.md5
                    WHERE c.relative_path=? AND c.md5=?
                    """,
                    (image.relative_path, file_md5),
                ).fetchone()
            finally:
                conn.close()
        if row is None:
            return None
        return {"has_text": row[0], "test_str": row[1], "operation": row[2]}

    def detect(image: ImageFile) -> bool:
        file_md5 = _file_md5(image.full_path)
        entry = _entry_for(image, file_md5)
        if entry is not None and entry.get("has_text") is not None:
            return bool(entry.get("has_text"))

        text = run_ocr(image)
        has_text = bool(re.search(r"[\u4e00-\u9fff]", text or ""))
        operation = entry.get("operation") if entry is not None else None
        now_text = _now_text()
        with db_lock:
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """
                    INSERT INTO ocr_cache(relative_path, md5, has_text, test_str, operation, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(relative_path) DO UPDATE SET
                        md5=excluded.md5,
                        has_text=excluded.has_text,
                        test_str=excluded.test_str,
                        operation=excluded.operation,
                        updated_at=excluded.updated_at
                    """,
                    (
                        image.relative_path,
                        file_md5,
                        int(has_text),
                        text or "",
                        operation if isinstance(operation, str) else None,
                        now_text,
                    ),
                )
                conn.commit()
            finally:
                conn.close()
        return has_text

    def ocr_text_for(image: ImageFile) -> str | None:
        file_md5 = _file_md5(image.full_path)
        entry = _entry_for(image, file_md5)
        text = entry.get("test_str") if entry is not None else None
        return text if isinstance(text, str) and text else None

    def operation_for(image: ImageFile) -> str | None:
        file_md5 = _file_md5(image.full_path)
        entry = _entry_for(image, file_md5)
        operation = entry.get("operation") if entry is not None else None
        return operation if isinstance(operation, str) else None

    def md5_for(image: ImageFile) -> str | None:
        return _file_md5(image.full_path)

    setattr(detect, "ocr_text_for", ocr_text_for)
    setattr(detect, "operation_for", operation_for)
    setattr(detect, "md5_for", md5_for)
    return detect


def ocr_text_detector_factory(
    cache_path: Path | None = None,
) -> Callable[[ImageFile], bool]:
    cache: dict[str, dict[str, object]] = {}
    if cache_path is not None:
        cache = load_ocr_cache(cache_path)
    cache_lock = threading.Lock()

    def detect(image: ImageFile) -> bool:
        file_key = image.relative_path
        file_md5 = _file_md5(image.full_path)
        setattr(detect, "_last_md5", file_md5)
        setattr(detect, "_last_operation", None)
        with cache_lock:
            entry = cache.get(file_key)
            if (
                isinstance(entry, dict)
                and entry.get("md5") == file_md5
                and "has_text" in entry
            ):
                operation = entry.get("operation")
                if isinstance(operation, str):
                    setattr(detect, "_last_operation", operation)
                if "has_test" in entry:
                    entry.pop("has_test", None)
                    if cache_path is not None:
                        save_ocr_cache(cache_path, cache)
                return bool(entry.get("has_text", False))

        text = run_ocr(image)
        has_text = bool(re.search(r"[\u4e00-\u9fff]", text or ""))
        with cache_lock:
            existing_entry = cache.get(file_key)
            operation = (
                existing_entry.get("operation")
                if isinstance(existing_entry, dict) and existing_entry.get("md5") == file_md5
                else None
            )
            new_entry: dict[str, object] = {
                "md5": file_md5,
                "has_text": has_text,
                "test_str": text or "",
            }
            if isinstance(operation, str):
                new_entry["operation"] = operation
            cache[file_key] = new_entry
            if cache_path is not None:
                save_ocr_cache(cache_path, cache)
        return has_text

    def ocr_text_for(image: ImageFile) -> str | None:
        with cache_lock:
            entry = cache.get(image.relative_path)
        if isinstance(entry, dict):
            text = entry.get("test_str")
            if isinstance(text, str) and text:
                return text
        return None

    def operation_for(image: ImageFile) -> str | None:
        file_md5 = _file_md5(image.full_path)
        with cache_lock:
            entry = cache.get(image.relative_path)
        if isinstance(entry, dict) and entry.get("md5") == file_md5:
            operation = entry.get("operation")
            return operation if isinstance(operation, str) else None
        return None

    def md5_for(image: ImageFile) -> str | None:
        return _file_md5(image.full_path)

    setattr(detect, "ocr_text_for", ocr_text_for)
    setattr(detect, "operation_for", operation_for)
    setattr(detect, "md5_for", md5_for)
    return detect

_TRANSLATION_SYSTEM_PROMPT = (
    "你是游戏本地化翻译质量检查专家，"
    "专注于中文游戏《剑厑3》的简体→繁体转换。\n"
    "你需要判断：国际版（繁体中文）图片中的文字，"
    "是否是对应陆版（简体中文）文字的准确繁体翻译。\n\n"
    "判断标准：\n"
    "1. 繁体字形是否正确（不得残留简体字）\n"
    "2. 内容是否完整（无漏译）\n"
    "3. 语义是否准确（无错译、无语法错误）\n\n"
    "若图片为 UI 元素（按鈕、标签、提示），以游戏术语习惯为准。\n\n"
    "请严格按以下格式回复（不要有多余内容）：\n"
    "状态：OK\n"
    "或\n"
    "状态：ERROR\n"
    "说明：<具体问题，一句话>"
)


def make_translation_checker(
    api_key: str,
) -> Callable[[str, str | None], tuple[str, str]]:
    """
    Return a translation check function: check(mainland_text, i18n_text) -> (status, note).
    status values: "ok" | "error" | "missing"
    """
    try:
        import anthropic
    except ImportError:
        raise SystemExit(
            "ERROR: 翻译检查需要 anthropic 库，请先执行：pip install anthropic"
        )

    client = anthropic.Anthropic(api_key=api_key)

    def check(mainland_text: str, i18n_text: str | None) -> tuple[str, str]:
        if not i18n_text or not i18n_text.strip():
            return "missing", "国际版图片未识别到文字，可能缺少翻译"

        try:
            response = client.messages.create(
                model="claude-opus-4-7",
                max_tokens=300,
                system=[
                    {
                        "type": "text",
                        "text": _TRANSLATION_SYSTEM_PROMPT,
                        "cache_control": {"type": "ephemeral"},
                    }
                ],
                messages=[
                    {
                        "role": "user",
                        "content": (
                            f"陆版文字：\n{mainland_text}\n\n"
                            f"国际版文字：\n{i18n_text}"
                        ),
                    }
                ],
            )
            reply = response.content[0].text.strip()
        except Exception as exc:
            print(f"WARN: Claude API 调用失败: {exc}", file=sys.stderr)
            return "error", f"API 调用失败: {exc}"

        if "状态：OK" in reply or "状态:OK" in reply:
            return "ok", ""
        note = ""
        for line in reply.splitlines():
            stripped = line.strip()
            if stripped.startswith("说明：") or stripped.startswith("说明:"):
                note = stripped.split("：", 1)[-1].split(":", 1)[-1].strip()
                break
        return "error", note or reply

    return check


def enrich_findings_with_translation(
    findings: list[Finding],
    translation_checker: Callable[[str, str | None], tuple[str, str]],
) -> list[Finding]:
    """
    For mainland_changed / mainland_new_with_text findings, run OCR on both images
    and call translation_checker to assess translation quality.
    Returns a new list; original findings are not mutated.
    """
    NEEDS_CHECK = {"mainland_changed", "mainland_new_with_text"}
    result: list[Finding] = []
    for f in findings:
        if f.issue not in NEEDS_CHECK:
            result.append(f)
            continue

        mainland_text = run_ocr_from_path(f.mainland_path, f.category)
        if not mainland_text or not re.search(r"[一-鿿]", mainland_text):
            result.append(dataclasses.replace(f, translation_status="no_cn_text"))
            continue

        i18n_text = run_ocr_from_path(f.i18n_path, f.category) if f.i18n_path else None
        status, note = translation_checker(mainland_text, i18n_text)
        result.append(
            dataclasses.replace(
                f,
                mainland_ocr_text=mainland_text,
                i18n_ocr_text=i18n_text,
                translation_status=status,
                translation_note=note,
            )
        )
    return result


def load_last_check(path: Path) -> dt.datetime | None:
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return None
    return dt.datetime.fromisoformat(text.replace("Z", "+00:00"))


def save_last_check(path: Path, value: dt.datetime) -> None:
    path.write_text(value.astimezone(dt.timezone.utc).isoformat(), encoding="utf-8")


def xlsx_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        value = value.astimezone(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")
    return html.escape(str(value), quote=True)


def cell_name(col_index: int, row_index: int) -> str:
    col = ""
    n = col_index
    while n:
        n, rem = divmod(n - 1, 26)
        col = chr(65 + rem) + col
    return f"{col}{row_index}"


def write_xlsx(
    path: Path,
    findings: Sequence[Finding],
    thumbnail_limit: int | None = None,
    thumbnail_issues: set[str] | None = None,
    no_thumbnail_resize: bool = False,
    max_image_px: int | None = None,
    stats: dict[str, int] | None = None,
) -> None:
    if stats is None:
        stats = {}
    if thumbnail_issues is None:
        thumbnail_issues = DEFAULT_THUMBNAIL_ISSUES
    try:
        write_xlsx_with_openpyxl(
            path,
            findings,
            thumbnail_limit,
            thumbnail_issues,
            no_thumbnail_resize,
            max_image_px,
            stats,
        )
        return
    except ImportError:
        pass
    except Exception as exc:
        print(f"WARN: 带缩略图 Excel 输出失败，降级为纯文本 xlsx: {exc}", file=sys.stderr)

    rows: list[list[object]] = [
        [
            "类别",
            "问题类型",
            "相对路径",
            "国际版路径",
            "陆版路径",
            "国际版缩略图",
            "陆版缩略图",
            "国际版修改时间",
            "陆版修改时间",
            "说明",
            "陆版识别文字",
            "国际版识别文字",
            "翻译状态",
            "翻译问题说明",
        ]
    ]
    for finding in findings:
        rows.append(
            [
                finding.category,
                finding.issue,
                finding.relative_path,
                finding.i18n_path,
                finding.mainland_path,
                "",
                "",
                finding.i18n_modified_at,
                finding.mainland_modified_at,
                finding.detail,
                finding.mainland_ocr_text,
                finding.i18n_ocr_text,
                finding.translation_status,
                finding.translation_note,
            ]
        )

    sheet_rows: list[str] = []
    for row_index, row in enumerate(rows, 1):
        cells = []
        for col_index, value in enumerate(row, 1):
            ref = cell_name(col_index, row_index)
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{xlsx_text(value)}</t></is></c>')
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData></worksheet>'
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="检查结果" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )
    wb_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/></Relationships>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def write_xlsx_with_openpyxl(
    path: Path,
    findings: Sequence[Finding],
    thumbnail_limit: int | None = None,
    thumbnail_issues: set[str] | None = None,
    no_thumbnail_resize: bool = False,
    max_image_px: int | None = None,
    stats: dict[str, int] | None = None,
) -> None:
    if stats is None:
        stats = {}
    if thumbnail_issues is None:
        thumbnail_issues = DEFAULT_THUMBNAIL_ISSUES
    if no_thumbnail_resize:
        box: tuple[int, int] | None = None
    elif max_image_px is not None and max_image_px > 0:
        box = (max_image_px, max_image_px)
    else:
        box = (96, 72)
    dynamic_row_height = no_thumbnail_resize or (
        max_image_px is not None and max_image_px > 0
    )
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    wb = Workbook()
    ws = wb.active
    ws.title = "检查结果"

    # ---- summary rows ----
    missing_count = sum(1 for f in findings if f.issue == "mainland_missing")
    changed_count = sum(1 for f in findings if f.issue == "mainland_changed")
    new_text_count = sum(1 for f in findings if f.issue == "mainland_new_with_text")
    other_count = len(findings) - missing_count - changed_count - new_text_count
    normal_synced = stats.get("normal_synced", 0)
    new_no_text = stats.get("new_no_text", 0)

    summary_header_font = Font(bold=True, size=13)
    summary_rows = [
        ["一、共检查图片", "", ""],
        ["国际版图片", f"{stats.get('i18n_count', 0)} 张", ""],
        ["陆版图片", f"{stats.get('mainland_count', 0)} 张", ""],
        ["", "", ""],
        ["二、正常图片", f"{normal_synced + new_no_text} 张", ""],
        ["正常同步文件", f"{normal_synced} 张", "两边都有且修改时间正常"],
        ["新增无文字图片", f"{new_no_text} 张", "陆版新增但无文字，无需翻译"],
        ["", "", ""],
        ["三、异常图片", f"{len(findings)} 张", ""],
    ]
    if missing_count:
        summary_rows.append(["疑似废除文件", f"{missing_count} 张", "国际版有、陆版无（可能已废弃）"])
    if changed_count:
        summary_rows.append(["修改时间异常", f"{changed_count} 张", "陆版修改时间晚于国际版（需更新翻译）"])
    if new_text_count:
        summary_rows.append(["新增带文字图片", f"{new_text_count} 张", "陆版新增含文字图片（缺少国际版对应翻译）"])
    if other_count:
        summary_rows.append(["其他问题", f"{other_count} 张", ""])
    summary_rows.append(["", "", ""])

    for row_idx, row_data in enumerate(summary_rows, 1):
        ws.append(row_data)
        if row_idx == 1:
            ws[row_idx][0].font = summary_header_font
        elif row_data[0].startswith("二、") or row_data[0].startswith("三、"):
            ws[row_idx][0].font = Font(bold=True)

    # ---- data header ----
    headers = [
        "类别",
        "问题类型",
        "相对路径",
        "国际版路径",
        "陆版路径",
        "国际版缩略图",
        "陆版缩略图",
        "国际版修改时间",
        "陆版修改时间",
        "说明",
        "陆版识别文字",
        "国际版识别文字",
        "翻译状态",
        "翻译问题说明",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thumb_dir = path.parent / f".{path.stem}_thumbs"
    thumb_dir.mkdir(parents=True, exist_ok=True)
    temp_files: list[Path] = []

    inserted_thumbnails = 0
    max_col_px = {"F": 0, "G": 0}
    try:
        data_start = len(summary_rows) + 2  # summary + 1 header row
        for row_index, finding in enumerate(findings, data_start):
            ws.append(
                [
                    finding.category,
                    finding.issue,
                    finding.relative_path,
                    finding.i18n_path,
                    finding.mainland_path,
                    "",
                    "",
                    format_dt(finding.i18n_modified_at),
                    format_dt(finding.mainland_modified_at),
                    finding.detail,
                    finding.mainland_ocr_text or "",
                    finding.i18n_ocr_text or "",
                    finding.translation_status or "",
                    finding.translation_note or "",
                ]
            )
            ws.row_dimensions[row_index].height = 72
            thumbnail_keys = {finding.issue}
            if finding.detail:
                thumbnail_keys.add("__has_detail__")
            if finding.translation_status:
                thumbnail_keys.add(f"translation_{finding.translation_status}")
            should_embed = bool(thumbnail_keys & thumbnail_issues)
            under_limit = thumbnail_limit is None or inserted_thumbnails < thumbnail_limit
            max_px_h = 0
            if should_embed and under_limit:
                inserted, px_w, px_h = add_thumbnail(
                    ws,
                    PILImage,
                    XlsxImage,
                    finding.i18n_path,
                    f"F{row_index}",
                    thumb_dir,
                    temp_files,
                    no_resize=no_thumbnail_resize,
                    max_box=box,
                )
                inserted_thumbnails += inserted
                max_px_h = max(max_px_h, px_h)
                max_col_px["F"] = max(max_col_px["F"], px_w)
                if thumbnail_limit is None or inserted_thumbnails < thumbnail_limit:
                    inserted, px_w, px_h = add_thumbnail(
                        ws,
                        PILImage,
                        XlsxImage,
                        finding.mainland_path,
                        f"G{row_index}",
                        thumb_dir,
                        temp_files,
                        no_resize=no_thumbnail_resize,
                        max_box=box,
                    )
                    inserted_thumbnails += inserted
                    max_px_h = max(max_px_h, px_h)
                    max_col_px["G"] = max(max_col_px["G"], px_w)
            if dynamic_row_height and max_px_h > 0:
                ws.row_dimensions[row_index].height = max(72, max_px_h * 0.75)

        widths = [14, 22, 48, 70, 70, 14, 14, 24, 24, 36, 40, 40, 14, 50]
        if dynamic_row_height:
            for col_letter, idx in (("F", 5), ("G", 6)):
                if max_col_px[col_letter] > 0:
                    widths[idx] = max(widths[idx], max_col_px[col_letter] / 7 + 2)
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = f"A{data_start}"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(path)
    finally:
        for file in temp_files:
            try:
                file.unlink()
            except OSError:
                pass
        try:
            thumb_dir.rmdir()
        except OSError:
            pass


def format_dt(value: dt.datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")


def issue_title(issue: str) -> str:
    return {
        "mainland_missing": TEXT_MISSING_ISSUE,
        "mainland_changed": TEXT_CHANGED_ISSUE,
        "mainland_new_with_text": TEXT_NEW_TEXT_ISSUE,
        "mainland_new_no_text": TEXT_NEW_NO_TEXT_ISSUE,
    }.get(issue, issue)


def file_uri(path: str) -> str:
    if not path:
        return ""
    if path.lower().startswith(("http://", "https://")):
        return path
    return Path(path).resolve().as_uri()


def make_html_image_asset(
    source: str,
    assets_dir: Path,
    index: int,
    side: str,
    max_image_px: int | None = None,
) -> str:
    local = local_image_path(source)
    if local is None:
        return file_uri(source)
    assets_dir.mkdir(parents=True, exist_ok=True)
    out = assets_dir / f"img_{index}_{side}{local.suffix.lower()}.png"

    def write_unavailable_placeholder(reason: str) -> str:
        try:
            from PIL import Image as PILImage
            from PIL import ImageDraw

            img = PILImage.new("RGB", (360, 120), (248, 250, 252))
            draw = ImageDraw.Draw(img)
            draw.rectangle((0, 0, 359, 119), outline=(203, 213, 225))
            draw.text((16, 24), "Preview unavailable", fill=(51, 65, 85))
            draw.text((16, 52), local.name[:42], fill=(100, 116, 139))
            draw.text((16, 80), reason[:42], fill=(100, 116, 139))
            img.save(out, format="PNG")
        except Exception:
            # Last-resort PNG so the HTML never points at browser-unsupported .tga/.dds files.
            import struct
            import zlib

            width, height = 360, 120
            row = b"\x00" + bytes((248, 250, 252)) * width
            raw = row * height

            def chunk(kind: bytes, payload: bytes) -> bytes:
                return (
                    struct.pack(">I", len(payload))
                    + kind
                    + payload
                    + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)
                )

            png = (
                b"\x89PNG\r\n\x1a\n"
                + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
                + chunk(b"IDAT", zlib.compress(raw))
                + chunk(b"IEND", b"")
            )
            out.write_bytes(png)
        return f"{assets_dir.name}/{out.name}"

    try:
        from PIL import Image as PILImage

        with PILImage.open(local) as img:
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            max_px = max_image_px if max_image_px is not None and max_image_px > 0 else 720
            img.thumbnail((max_px, max_px))
            img.save(out, format="PNG")
        return f"{assets_dir.name}/{out.name}"
    except Exception as exc:
        print(f"WARN: 无法生成 HTML 预览图，使用占位图: {local}: {exc}", file=sys.stderr)
        try:
            return write_unavailable_placeholder(str(exc))
        except Exception as placeholder_exc:
            print(f"WARN: 无法生成 HTML 占位图: {local}: {placeholder_exc}", file=sys.stderr)
            return file_uri(source)


def html_img(asset: str, label: str) -> str:
    if not asset:
        return ""
    escaped_asset = html.escape(asset, quote=True)
    escaped_label = html.escape(label, quote=True)
    return (
        f'<button class="thumb-button" type="button" onclick="openPreview(this)" '
        f'data-full-src="{escaped_asset}" data-title="{escaped_label}">'
        f'<img src="{escaped_asset}" alt="{escaped_label}" loading="lazy"></button>'
    )


def build_report_title(now: dt.datetime | None = None) -> str:
    current = now or dt.datetime.now(BEIJING_TZ)
    if current.tzinfo is None:
        current = current.replace(tzinfo=BEIJING_TZ)
    current = current.astimezone(BEIJING_TZ)
    return f"\u300a{TEXT_PROJECT_NAME}\u300b{TEXT_REPORT_TITLE}\uff08{current:%Y.%m.%d}\uff09"


def write_html_report(
    path: Path,
    findings: Sequence[Finding],
    i18n_count: int = 0,
    mainland_count: int = 0,
    normal_synced: int = 0,
    new_no_text: int = 0,
    max_image_px: int | None = None,
    ocr_cache_name: str = OCR_CACHE_FILE,
) -> None:
    assets_dir = path.parent / f"{path.stem}_assets"
    if assets_dir.exists():
        shutil.rmtree(assets_dir)
    missing = [f for f in findings if f.issue == "mainland_missing"]
    changed = [f for f in findings if f.issue == "mainland_changed"]
    new_with_text = [f for f in findings if f.issue == "mainland_new_with_text"]
    new_without_text = [f for f in findings if f.issue == "mainland_new_no_text"]
    others = [
        f for f in findings
        if f.issue not in {
            "mainland_missing",
            "mainland_changed",
            "mainland_new_with_text",
            "mainland_new_no_text",
        }
    ]

    issue_type_labels = {
        "mainland_missing": "\u56fd\u9645\u7248\u5b58\u5728\u4f46\u9646\u7248\u7f3a\u5931\uff0c\u7591\u4f3c\u5e9f\u5f03\u6216\u8def\u5f84\u4e0d\u4e00\u81f4",
        "mainland_changed": "\u9646\u7248\u4fee\u6539\u65f6\u95f4\u665a\u4e8e\u56fd\u9645\u7248\uff0c\u9700\u8981\u786e\u8ba4\u56fd\u9645\u7248\u662f\u5426\u540c\u6b65",
        "mainland_new_with_text": "\u9646\u7248\u65b0\u589e\u4e14 OCR \u8bc6\u522b\u5230\u6587\u5b57\uff0c\u9700\u8981\u8865\u5145\u56fd\u9645\u7248\u7ffb\u8bd1",
        "mainland_new_no_text": "\u9646\u7248\u65b0\u589e\u4f46 OCR \u672a\u8bc6\u522b\u6587\u5b57\uff0c\u8fdb\u5165\u8be6\u60c5\u4f9b\u4eba\u5de5\u590d\u6838",
    }

    def issue_summary_card(issue_key: str, items: list[Finding], index: int) -> str:
        label = issue_type_labels.get(issue_key, issue_key)
        return (
            f'<section class="summary-card {html.escape(issue_key)}">'
            f'<h3>{index}. {html.escape(issue_title(issue_key))}: {len(items)} {TEXT_ITEM}</h3>'
            f'<p>{html.escape(label)}</p>'
            f'<ul>{_summary_items(items)}</ul></section>'
        )

    summary_cards: list[str] = []
    card_index = 0
    for issue_key, items in (
        ("mainland_missing", missing),
        ("mainland_changed", changed),
        ("mainland_new_with_text", new_with_text),
        ("mainland_new_no_text", new_without_text),
    ):
        if items:
            card_index += 1
            summary_cards.append(issue_summary_card(issue_key, items, card_index))
    if others:
        card_index += 1
        summary_cards.append(
            f'<section class="summary-card other"><h3>{card_index}. {TEXT_OTHER_ISSUE}: {len(others)} {TEXT_ITEM}</h3>'
            f'<ul>{_summary_items(others)}</ul></section>'
        )

    detail_findings = [f for f in findings if f.issue != "mainland_new_no_text"]
    rows = []
    for index, finding in enumerate(detail_findings, 1):
        i18n_asset = make_html_image_asset(finding.i18n_path, assets_dir, index, "i18n", max_image_px)
        mainland_asset = make_html_image_asset(finding.mainland_path, assets_dir, index, "mainland", max_image_px)
        issue_class = html.escape(finding.issue.replace("_", "-"))
        ocr_text = finding.mainland_ocr_text or ""
        row_title = ""
        if finding.issue == "mainland_new_with_text" and ocr_text:
            row_title = f"{TEXT_OCR_TEXT}\uff1a{ocr_text}"
        issue_label = issue_title(finding.issue)
        mainland_dt = format_dt(finding.mainland_modified_at)
        i18n_dt = format_dt(finding.i18n_modified_at)
        operation = finding.operation or ""
        can_ignore = finding.issue == "mainland_new_with_text" and bool(finding.mainland_md5)
        operation_label = TEXT_IGNORED if operation == OCR_OPERATION_IGNORE else TEXT_IGNORE
        operation_class = "ignored" if operation == OCR_OPERATION_IGNORE else ""
        if can_ignore:
            operation_cell = (
                f'<td><button class="operation-button {operation_class}" type="button" '
                f'onclick="ignoreFinding({index})">{html.escape(operation_label)}</button></td>'
            )
        else:
            operation_cell = "<td></td>"
        operation_filter_label = operation_label if can_ignore else ""
        rows.append({
            "rowIndex": index,
            "className": issue_class,
            "pairType": finding.category or TEXT_NONE,
            "title": row_title,
            "ocr": ocr_text,
            "operation": operation,
            "canIgnore": can_ignore,
            "relativePath": finding.relative_path,
            "mainlandMd5": finding.mainland_md5 or "",
            "cells": [
                f'<td class="row-index">{index}</td>',
                f'<td><span class="issue-badge {issue_class}">{html.escape(issue_label)}</span></td>',
                f'<td class="path">{html.escape(finding.relative_path)}</td>',
                f'<td>{html_img(mainland_asset, TEXT_MAINLAND + " " + finding.relative_path)}</td>',
                f'<td>{html_img(i18n_asset, TEXT_I18N + " " + finding.relative_path)}</td>',
                f'<td>{html.escape(mainland_dt)}</td>',
                f'<td>{html.escape(i18n_dt)}</td>',
                f'<td>{html.escape(finding.detail)}</td>',
                operation_cell,
            ],
            "filterValues": [
                str(index),
                issue_label,
                finding.relative_path,
                "",
                "",
                mainland_dt,
                i18n_dt,
                finding.detail,
                operation_filter_label,
            ],
        })

    total_abnormal = len(findings) - len(new_without_text)
    report_title = build_report_title()
    summary_table = _build_summary_table(
        i18n_count, mainland_count, total_abnormal,
        normal_synced, new_no_text,
        missing, changed, new_with_text, others, new_without_text, report_title,
    )
    detail_types = sorted({f.category or TEXT_NONE for f in detail_findings}) or [TEXT_NONE]
    tab_buttons = "".join(
        f'<button type="button" class="tab-button" data-tab-type="{html.escape(pair_type, quote=True)}" '
        f'onclick="switchTab(this.dataset.tabType)">{html.escape(pair_type)}</button>'
        for pair_type in detail_types
    )
    doc = _html_template(
        summary_table,
        summary_cards,
        rows,
        tab_buttons,
        detail_types[0],
        report_title,
        ocr_cache_name,
    )
    path.write_text(doc, encoding="utf-8")

def _summary_items(items: Sequence[Finding]) -> str:
    if not items:
        return f"<li>{TEXT_NONE}</li>"
    return "".join(
        f"<li>{html.escape(f.relative_path)} - {html.escape(f.detail)}</li>"
        for f in items
    )


def _build_summary_table(
    i18n_count: int,
    mainland_count: int,
    total_abnormal: int,
    normal_synced: int,
    new_no_text: int,
    missing: list[Finding],
    changed: list[Finding],
    new_with_text: list[Finding],
    others: list[Finding],
    new_without_text: list[Finding] | None = None,
    report_title: str = TEXT_REPORT_TITLE,
) -> str:
    new_without_text = new_without_text or []
    checked_rows = "".join(
        f"<tr><td>{label}</td><td>{count} 张</td></tr>"
        for label, count in [
            ("\u56fd\u9645\u7248\u56fe\u7247", i18n_count),
            ("\u9646\u7248\u56fe\u7247", mainland_count),
        ]
    )
    normal_rows = "".join(
        f"<tr><td>{label}</td><td>{count} 张</td><td>{note}</td></tr>"
        for label, count, note in [
            ("\u6b63\u5e38\u540c\u6b65\u6587\u4ef6", normal_synced, "\u4e24\u8fb9\u90fd\u6709\u4e14\u4fee\u6539\u65f6\u95f4\u6b63\u5e38"),
            ("\u65b0\u589e\u65e0\u6587\u5b57\u56fe\u7247", len(new_without_text), "\u9646\u7248\u65b0\u589e\u65e0\u6587\u5b57\uff0c\u8be6\u60c5\u4e2d\u4f9b\u4eba\u5de5\u590d\u6838"),
        ]
    )
    abnormal_rows = "".join(
        f"<tr><td>{label}</td><td>{count} 张</td><td>{note}</td></tr>"
        for label, count, note in [
            ("\u7591\u4f3c\u5e9f\u5f03\u6587\u4ef6", len(missing), "\u56fd\u9645\u7248\u6709\u3001\u9646\u7248\u65e0\uff08\u53ef\u80fd\u5df2\u5e9f\u5f03\uff09"),
            ("\u4fee\u6539\u65f6\u95f4\u5f02\u5e38", len(changed), "\u9646\u7248\u4fee\u6539\u65f6\u95f4\u665a\u4e8e\u56fd\u9645\u7248\uff08\u9700\u66f4\u65b0\u7ffb\u8bd1\uff09"),
            ("\u65b0\u589e\u5e26\u6587\u5b57\u56fe\u7247", len(new_with_text), "\u9646\u7248\u65b0\u589e\u542b\u6587\u5b57\u56fe\u7247\uff08\u7f3a\u5c11\u56fd\u9645\u7248\u5bf9\u5e94\u7ffb\u8bd1\uff09"),
        ]
    )
    if others:
        abnormal_rows += "".join(
            f"<tr><td>{html.escape(issue_title(f.issue))}</td><td>1 张</td><td>{html.escape(f.detail)}</td></tr>"
            for f in others
        )
    return f"""
    <section class="report-shell">
      <div class="hero-panel">
        <div>
          <p class="eyebrow">Image Localization Audit</p>
          <h1>{html.escape(report_title)}</h1>
          <p class="subtitle">\u81ea\u52a8\u68c0\u67e5\u56fd\u9645\u7248\u4e0e\u9646\u7248\u56fe\u7247\u5dee\u5f02\uff0c\u8f93\u51fa\u5f02\u5e38\u9879\u3001\u590d\u6838\u9879\u3001\u7f29\u7565\u56fe\u548c OCR \u8bc6\u522b\u6587\u672c\u3002</p>
        </div>
        <div class="run-result"><span>{TEXT_RUN_RESULT}</span><strong>{total_abnormal}</strong><small>\u5f02\u5e38\u9879</small></div>
      </div>
      <div class="issue-breakdown">
        <p><strong>\u4e00\u3001\u5171\u68c0\u67e5\u56fe\u7247</strong></p>
        <table class="summary-table"><thead><tr><th>\u6307\u6807</th><th>\u6570\u91cf</th></tr></thead><tbody>{checked_rows}</tbody></table>
        <p class="summary-heading"><strong>\u4e8c\u3001\u6b63\u5e38\u56fe\u7247\uff1a\u5171 {normal_synced + len(new_without_text)} \u5f20</strong></p>
        <table class="summary-table"><thead><tr><th>\u6b63\u5e38\u7c7b\u578b</th><th>\u6570\u91cf</th><th>\u8bf4\u660e</th></tr></thead><tbody>{normal_rows}</tbody></table>
        <p class="summary-heading"><strong>\u4e09\u3001\u5f02\u5e38\u56fe\u7247\uff1a\u5171 <span class="abnormal-count">{total_abnormal}</span> \u5f20</strong></p>
        <table class="summary-table"><thead><tr><th>\u95ee\u9898\u7c7b\u578b</th><th>\u6570\u91cf</th><th>\u8bf4\u660e</th></tr></thead><tbody>{abnormal_rows}</tbody></table>
      </div>
    </section>
    """


def _html_template(
    summary_table: str,
    summary_cards: list[str],
    rows: list[dict[str, object]],
    tab_buttons: str = "",
    default_tab_type: str = TEXT_NONE,
    report_title: str = TEXT_REPORT_TITLE,
    ocr_cache_name: str = OCR_CACHE_FILE,
) -> str:
    rows_json = json.dumps(rows, ensure_ascii=False)
    default_tab_json = json.dumps(default_tab_type, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>{html.escape(report_title)}</title>
<style>
:root {{ --bg:#f4f6f8; --panel:#ffffff; --text:#1f2937; --muted:#64748b; --line:#d8dee8; --header:#102033; --danger:#b42318; --warning:#b76e00; --review:#475569; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:"Microsoft YaHei", "Segoe UI", Arial, sans-serif; color:var(--text); background:var(--bg); }}
main {{ max-width:1680px; margin:0 auto; padding:28px; }}
.report-shell {{ display:grid; gap:18px; }}
.hero-panel {{ display:flex; justify-content:space-between; gap:24px; align-items:stretch; background:linear-gradient(135deg,#102033,#243b55); color:white; border-radius:8px; padding:26px 30px; box-shadow:0 14px 30px rgba(15,23,42,.14); }}
.hero-panel h1 {{ margin:4px 0 8px; font-size:30px; letter-spacing:0; }}
.eyebrow {{ margin:0; color:#c6d3e1; font-size:12px; text-transform:uppercase; letter-spacing:.08em; }}
.subtitle {{ margin:0; color:#d9e3ef; font-size:14px; }}
.run-result {{ min-width:160px; border:1px solid rgba(255,255,255,.24); border-radius:6px; padding:16px; text-align:center; background:rgba(255,255,255,.08); }}
.run-result span,.run-result small {{ display:block; color:#d9e3ef; }}
.run-result strong {{ display:block; font-size:42px; line-height:1.1; }}
.metric-card,.issue-breakdown,.detail-panel {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; }}
.metric-card {{ padding:14px 16px; }}
.metric-card span {{ display:block; color:var(--muted); font-size:13px; }}
.metric-card strong {{ display:block; margin-top:4px; font-size:28px; }}
.metric-card small {{ display:block; margin-top:6px; color:var(--muted); line-height:1.4; }}
.issue-breakdown {{ padding:16px 18px; }}
.issue-breakdown p {{ margin:6px 0; font-size:15px; line-height:1.8; }}
.summary-heading {{ margin-top:18px !important; }}
.abnormal-count {{ color:#d93025; font-weight:700; }}
h2 {{ margin:0 0 12px; font-size:18px; }}
.summary-table {{ width:100%; border-collapse:collapse; }}
.summary-table th,.summary-table td {{ border-bottom:1px solid #e7ebf1; padding:10px 12px; text-align:left; font-size:13px; }}
.summary-table th {{ background:#f0f4f8; color:#334155; font-weight:600; }}
.detail-panel {{ overflow:hidden; }}
.detail-toolbar {{ display:flex; align-items:flex-start; justify-content:space-between; gap:16px; padding:16px 18px; border-bottom:1px solid var(--line); }}
.detail-toolbar p {{ margin:0; color:var(--muted); font-size:13px; }}
.export-button {{ flex:0 0 auto; border:1px solid #2563eb; background:#2563eb; color:white; border-radius:5px; padding:8px 14px; font-size:13px; cursor:pointer; }}
.export-button:hover {{ background:#1d4ed8; }}
.detail-tabs {{ display:flex; gap:8px; flex-wrap:wrap; padding:12px 18px; border-bottom:1px solid var(--line); background:#fbfdff; }}
.tab-button {{ border:1px solid #cbd5e1; background:white; color:#334155; border-radius:5px; padding:7px 12px; font-size:13px; cursor:pointer; }}
.tab-button.active {{ border-color:#2563eb; background:#2563eb; color:white; }}
.pagination {{ display:flex; align-items:center; justify-content:flex-end; gap:8px; padding:12px 18px; border-top:1px solid var(--line); background:#fbfdff; }}
.pagination button {{ border:1px solid #cbd5e1; background:white; color:#334155; border-radius:5px; padding:6px 10px; font-size:13px; cursor:pointer; }}
.pagination button:disabled {{ opacity:.45; cursor:not-allowed; }}
.pagination-status {{ color:var(--muted); font-size:13px; min-width:220px; text-align:center; }}
.table-wrap {{ overflow:auto; max-height:calc(100vh - 160px); }}
table {{ width:100%; min-width:1180px; border-collapse:separate; border-spacing:0; table-layout:fixed; }}
th,td {{ border-bottom:1px solid #e7ebf1; padding:10px; vertical-align:top; font-size:13px; }}
th {{ position:sticky; top:0; height:40px; background:#eef3f8; color:#334155; z-index:3; text-align:left; cursor:pointer; user-select:none; white-space:nowrap; }}
.filter-row th {{ top:40px; background:#f8fafc; cursor:default; z-index:2; }}
.filter-row input,.filter-row select {{ width:100%; padding:6px 8px; border:1px solid #cbd5e1; border-radius:4px; font-size:12px; background:white; }}
.date-range {{ display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:4px; }}
.date-range span {{ color:var(--muted); font-size:12px; }}
.date-range input {{ min-width:0; padding:6px 4px; }}
.row-index {{ text-align:center; color:#475569; }}
.path {{ word-break:break-all; }}
tr.mainland-new-with-text[title] {{ cursor:help; }}
.issue-badge {{ display:inline-flex; align-items:center; border-radius:999px; padding:3px 8px; font-size:12px; font-weight:600; white-space:nowrap; }}
.issue-badge.mainland-missing,.issue-badge.mainland-new-with-text {{ color:#7a271a; background:#fee4e2; }}
.issue-badge.mainland-changed {{ color:#7a4a00; background:#ffefd0; }}
.issue-badge.mainland-new-no-text {{ color:#334155; background:#e2e8f0; }}
.thumb-button {{ width:168px; height:110px; padding:0; border:1px solid #cbd5e1; border-radius:5px; background:#f8fafc; cursor:zoom-in; display:inline-flex; align-items:center; justify-content:center; }}
.thumb-button img {{ max-width:166px; max-height:108px; object-fit:contain; }}
.operation-button {{ border:1px solid #cbd5e1; background:white; color:#334155; border-radius:5px; padding:6px 10px; font-size:13px; cursor:pointer; white-space:nowrap; }}
.operation-button:hover {{ border-color:#2563eb; color:#1d4ed8; }}
.operation-button.ignored {{ border-color:#94a3b8; background:#e2e8f0; color:#475569; cursor:default; }}
.overlay {{ position:fixed; inset:0; background:rgba(15,23,42,.82); display:none; align-items:center; justify-content:center; padding:32px; z-index:20; }}
.overlay.open {{ display:flex; }}
.preview {{ max-width:96vw; max-height:92vh; background:#fff; border-radius:8px; padding:12px; }}
.preview img {{ max-width:92vw; max-height:82vh; object-fit:contain; display:block; }}
.preview-title {{ margin:0 0 8px; font-size:14px; color:#334155; word-break:break-all; }}
.no-results {{ text-align:center; padding:40px; color:#94a3b8; font-size:15px; }}
</style>
</head>
<body>
<main>
{summary_table}
<section class="detail-panel">
  <div class="detail-toolbar"><div><h2>{TEXT_REPORT_DETAIL}</h2><p>详情仅展示需处理的异常图片；新增无文字图片列表已写入本次日志。</p></div><button class="export-button" type="button" onclick="exportFilteredRows()">\u5bfc\u51fa\u7b5b\u9009\u7ed3\u679c</button></div>
  <div class="detail-tabs" role="tablist">{tab_buttons}</div>
  <div class="table-wrap">
    <table id="detailTable">
      <colgroup>
        <col style="width:56px">
        <col style="width:150px">
        <col style="width:260px">
        <col style="width:180px">
        <col style="width:180px">
        <col style="width:170px">
        <col style="width:170px">
        <col style="width:220px">
        <col style="width:120px">
      </colgroup>
      <thead>
        <tr>
          <th onclick="sortTable(0)">\u5e8f\u53f7<span class="sort-icon"></span></th>
          <th onclick="sortTable(1)">\u95ee\u9898\u7c7b\u578b<span class="sort-icon"></span></th>
          <th onclick="sortTable(2)">\u76f8\u5bf9\u8def\u5f84<span class="sort-icon"></span></th>
          <th>\u9646\u7248\u56fe\u7247</th>
          <th>\u56fd\u9645\u7248\u56fe\u7247</th>
          <th onclick="sortTable(5)">\u9646\u7248\u4fee\u6539\u65f6\u95f4<span class="sort-icon"></span></th>
          <th onclick="sortTable(6)">\u56fd\u9645\u7248\u4fee\u6539\u65f6\u95f4<span class="sort-icon"></span></th>
          <th onclick="sortTable(7)">\u8bf4\u660e<span class="sort-icon"></span></th>
          <th onclick="sortTable(8)">{TEXT_OPERATION}<span class="sort-icon"></span></th>
        </tr>
        <tr class="filter-row">
          <th><input type="text" data-filter-col="0" oninput="filterTable()" placeholder="\u7b5b\u9009..."></th>
          <th><select data-filter-col="1" data-filter="issue" onchange="filterTable()"><option value="">\u5168\u90e8</option><option>{TEXT_MISSING_ISSUE}</option><option>{TEXT_CHANGED_ISSUE}</option><option>{TEXT_NEW_TEXT_ISSUE}</option><option>{TEXT_OTHER_ISSUE}</option></select></th>
          <th><input type="text" data-filter-col="2" oninput="filterTable()" placeholder="\u7b5b\u9009..."></th>
          <th></th><th></th>
          <th><div class="date-range"><input type="date" data-date-col="5" data-date-bound="start" onchange="filterTable()"><span>\u81f3</span><input type="date" data-date-col="5" data-date-bound="end" onchange="filterTable()"></div></th>
          <th><div class="date-range"><input type="date" data-date-col="6" data-date-bound="start" onchange="filterTable()"><span>\u81f3</span><input type="date" data-date-col="6" data-date-bound="end" onchange="filterTable()"></div></th>
          <th><input type="text" data-filter-col="7" oninput="filterTable()" placeholder="\u7b5b\u9009..."></th>
          <th><select data-filter-col="8" onchange="filterTable()"><option value="">\u5168\u90e8</option><option>{TEXT_IGNORE}</option><option>{TEXT_IGNORED}</option></select></th>
        </tr>
      </thead>
      <tbody></tbody>
    </table>
    <div class="no-results" id="noResults" style="display:none">\u6ca1\u6709\u5339\u914d\u7684\u7ed3\u679c</div>
  </div>
  <div class="pagination">
    <button type="button" onclick="gotoPage(1)" id="firstPageBtn">\u9996\u9875</button>
    <button type="button" onclick="gotoPage(currentPage - 1)" id="prevPageBtn">\u4e0a\u4e00\u9875</button>
    <span class="pagination-status" id="paginationStatus"></span>
    <button type="button" onclick="gotoPage(currentPage + 1)" id="nextPageBtn">\u4e0b\u4e00\u9875</button>
    <button type="button" onclick="gotoPage(totalPages)" id="lastPageBtn">\u672b\u9875</button>
  </div>
</section>
</main>
<div id="overlay" class="overlay" onclick="closePreview()"><div class="preview" onclick="event.stopPropagation()"><p id="previewTitle" class="preview-title"></p><img id="previewImage" alt=""></div></div>
<script>
const PAGE_SIZE = 50;
const allRows = {rows_json};
const FILTER_COLUMNS = [0, 1, 2, 7, 8];
const OCR_CACHE_FILE_NAME = '{html.escape(ocr_cache_name)}';
const OPERATION_IGNORE = '{OCR_OPERATION_IGNORE}';
const OCR_CACHE_OPERATION_API = '/api/ocr-cache/operation';
let activeTabType = {default_tab_json};
let filteredRows = allRows.filter(row => row.pairType === activeTabType);
let currentPage = 1;
let totalPages = 1;
let sortCol = -1;
let sortAsc = true;
allRows.forEach(row => updateRowOperation(row, row.operation || ''));
function switchTab(type) {{
  activeTabType = type;
  document.querySelectorAll('.tab-button').forEach(button => {{
    button.classList.toggle('active', button.dataset.tabType === activeTabType);
  }});
  filterTable();
}}
function sortTable(col) {{
  if (sortCol === col) sortAsc = !sortAsc; else {{ sortCol = col; sortAsc = true; }}
  filteredRows.sort((a,b) => {{
    let va = String(a.filterValues[col] || '').trim().toLowerCase();
    let vb = String(b.filterValues[col] || '').trim().toLowerCase();
    if (va < vb) return sortAsc ? -1 : 1;
    if (va > vb) return sortAsc ? 1 : -1;
    return 0;
  }});
  renderPage(1);
}}
function filterTable() {{
  const table = document.getElementById('detailTable');
  const controls = Array.from(table.querySelectorAll('.filter-row [data-filter-col]'))
    .filter(control => FILTER_COLUMNS.includes(Number(control.dataset.filterCol)));
  const dateRanges = collectDateRanges();
  filteredRows = allRows.filter(row => {{
    if (row.pairType !== activeTabType) return false;
    for (const control of controls) {{
      const filter = control.value.trim().toLowerCase();
      const columnIndex = Number(control.dataset.filterCol);
      const cellText = String(row.filterValues[columnIndex] || '').trim().toLowerCase();
      const isSelect = control.tagName === 'SELECT';
      if (filter && (isSelect ? cellText !== filter : !cellText.includes(filter))) return false;
    }}
    for (const range of dateRanges) {{
      if (!dateInRange(row.filterValues[range.columnIndex], range.start, range.end)) return false;
    }}
    return true;
  }});
  if (sortCol >= 0) {{
    const activeSortCol = sortCol;
    const activeSortAsc = sortAsc;
    filteredRows.sort((a,b) => {{
      let va = String(a.filterValues[activeSortCol] || '').trim().toLowerCase();
      let vb = String(b.filterValues[activeSortCol] || '').trim().toLowerCase();
      if (va < vb) return activeSortAsc ? -1 : 1;
      if (va > vb) return activeSortAsc ? 1 : -1;
      return 0;
    }});
  }}
  renderPage(1);
}}
function collectDateRanges() {{
  const ranges = new Map();
  document.querySelectorAll('.filter-row [data-date-col]').forEach(input => {{
    const columnIndex = Number(input.dataset.dateCol);
    const current = ranges.get(columnIndex) || {{ columnIndex, start: null, end: null }};
    current[input.dataset.dateBound] = parseDateOnly(input.value);
    ranges.set(columnIndex, current);
  }});
  return Array.from(ranges.values()).filter(range => range.start || range.end);
}}
function parseDateOnly(value) {{
  if (!value) return null;
  const parts = value.split('-').map(Number);
  if (parts.length !== 3 || parts.some(Number.isNaN)) return null;
  return new Date(parts[0], parts[1] - 1, parts[2]);
}}
function dateInRange(value, start, end) {{
  if (!start && !end) return true;
  const date = parseDateOnly(String(value || '').slice(0, 10));
  if (!date) return false;
  if (start && date < start) return false;
  if (end && date > end) return false;
  return true;
}}
function renderPage(page) {{
  totalPages = Math.max(1, Math.ceil(filteredRows.length / PAGE_SIZE));
  currentPage = Math.min(Math.max(1, page), totalPages);
  const start = (currentPage - 1) * PAGE_SIZE;
  const pageRows = filteredRows.slice(start, start + PAGE_SIZE);
  const tbody = document.querySelector('#detailTable tbody');
  tbody.innerHTML = pageRows.map(row => {{
    const title = row.title ? ` title="${{htmlEscape(row.title)}}"` : '';
    const ocr = row.ocr ? ` data-ocr="${{htmlEscape(row.ocr)}}"` : '';
    return `<tr class="${{htmlEscape(row.className)}}"${{title}}${{ocr}}>${{row.cells.join('')}}</tr>`;
  }}).join('');
  const noResults = filteredRows.length === 0;
  document.getElementById('noResults').style.display = noResults ? '' : 'none';
  document.getElementById('paginationStatus').textContent = noResults
    ? '\u65e0\u5339\u914d\u7ed3\u679c'
    : `\u7b2c ${{currentPage}} / ${{totalPages}} \u9875\uff0c\u5171 ${{filteredRows.length}} \u6761\uff0c\u6bcf\u9875 ${{PAGE_SIZE}} \u6761`;
  document.getElementById('firstPageBtn').disabled = currentPage <= 1 || noResults;
  document.getElementById('prevPageBtn').disabled = currentPage <= 1 || noResults;
  document.getElementById('nextPageBtn').disabled = currentPage >= totalPages || noResults;
  document.getElementById('lastPageBtn').disabled = currentPage >= totalPages || noResults;
}}
function gotoPage(page) {{
  renderPage(page);
}}
function operationStatus(operation) {{
  return operation === OPERATION_IGNORE ? '{TEXT_IGNORED}' : '{TEXT_IGNORE}';
}}
function buildOperationCell(row) {{
  if (!row.canIgnore) return '<td></td>';
  const ignored = row.operation === OPERATION_IGNORE;
  return `<td><button class="operation-button${{ignored ? ' ignored' : ''}}" type="button" onclick="ignoreFinding(${{row.rowIndex}})">${{ignored ? '{TEXT_IGNORED}' : '{TEXT_IGNORE}'}}</button></td>`;
}}
function updateRowOperation(row, operation) {{
  if (!row.canIgnore) {{
    row.operation = '';
    row.filterValues[8] = '';
    row.cells[8] = '<td></td>';
    return;
  }}
  row.operation = operation;
  row.filterValues[8] = operationStatus(operation);
  row.cells[8] = buildOperationCell(row);
}}
function findRowByIndex(rowIndex) {{
  return allRows.find(row => Number(row.rowIndex) === Number(rowIndex));
}}
async function readOcrCache() {{
  const response = await fetch(OCR_CACHE_FILE_NAME, {{ cache: 'no-store' }});
  if (!response.ok) return {{}};
  const text = await response.text();
  return text.trim() ? JSON.parse(text) : {{}};
}}
async function writeOcrCache(cache) {{
  const response = await fetch(OCR_CACHE_OPERATION_API, {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify(cache),
  }});
  if (!response.ok) {{
    throw new Error(`本地报告服务写入失败：${{response.status}}`);
  }}
}}
async function writeIgnoreOperation(row) {{
  try {{
    const response = await fetch(OCR_CACHE_OPERATION_API, {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{
        relativePath: row.relativePath,
        md5: row.mainlandMd5,
        operation: OPERATION_IGNORE,
      }}),
    }});
    if (response.ok) {{
      return;
    }}
  }} catch (error) {{
  }}
  const cache = await readOcrCache();
  const entry = cache[row.relativePath] && typeof cache[row.relativePath] === 'object'
    ? cache[row.relativePath]
    : {{}};
  entry.md5 = row.mainlandMd5;
  entry.operation = OPERATION_IGNORE;
  cache[row.relativePath] = entry;
  await writeOcrCache(cache);
}}
async function ignoreFinding(rowIndex) {{
  const row = findRowByIndex(rowIndex);
  if (!row || !row.canIgnore || row.operation === OPERATION_IGNORE) return;
  if (!row.relativePath || !row.mainlandMd5) {{
    alert('缺少缓存键或 md5，无法写入忽略标识。');
    return;
  }}
  try {{
    await writeIgnoreOperation(row);
    updateRowOperation(row, OPERATION_IGNORE);
    filterTable();
  }} catch (error) {{
    alert(`自动写入 ${{OCR_CACHE_FILE_NAME}} 失败：请不要用 file:// 直接打开，请用 --serve-report 启动的本地报告服务页面。${{error && error.message ? ' ' + error.message : ''}}`);
  }}
}}
function htmlEscape(value) {{
  return String(value || '')
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}}
async function imageSrcToDataUri(src) {{
  if (!src || src.startsWith('data:image')) return src || '';
  try {{
    const response = await fetch(src);
    if (!response.ok) return src;
    const blob = await response.blob();
    return await new Promise(resolve => {{
      const reader = new FileReader();
      reader.onload = () => resolve(reader.result || src);
      reader.onerror = () => resolve(src);
      reader.readAsDataURL(blob);
    }});
  }} catch (error) {{
    return src;
  }}
}}
async function buildExportImageCell(cell) {{
  const img = cell.querySelector('img');
  if (!img) return '';
  const src = img.getAttribute('src') || '';
  const alt = img.getAttribute('alt') || '';
  const dataSrc = await imageSrcToDataUri(src);
  return `<img src="${{htmlEscape(dataSrc)}}" alt="${{htmlEscape(alt)}}" style="max-width:120px;max-height:80px;">`;
}}
async function exportFilteredRows() {{
  const table = document.getElementById('detailTable');
  const headers = Array.from(table.querySelectorAll('thead tr:first-child th')).map(th => th.childNodes[0].textContent.trim());
  headers.push('{TEXT_OCR_TEXT}');
  const headerHtml = headers.map(header => `<th>${{htmlEscape(header)}}</th>`).join('');
  const rowHtml = await Promise.all(filteredRows.map(async row => {{
    const temp = document.createElement('tr');
    temp.innerHTML = row.cells.join('');
    const cellHtml = await Promise.all(Array.from(temp.cells).map(async (cell, index) => {{
      if (index === 3 || index === 4) return `<td>${{await buildExportImageCell(cell)}}</td>`;
      return `<td>${{htmlEscape(cell.textContent.trim())}}</td>`;
    }}));
    cellHtml.push(`<td>${{htmlEscape(row.ocr || '')}}</td>`);
    return `<tr>${{cellHtml.join('')}}</tr>`;
  }}));
  const bodyHtml = rowHtml.join('');
  const doc = `<!doctype html><html><head><meta charset="utf-8"><style>
    table {{ border-collapse:collapse; }}
    th,td {{ border:1px solid #999; padding:6px; vertical-align:top; mso-number-format:"\\@"; }}
    img {{ display:block; }}
  </style></head><body><table><thead><tr>${{headerHtml}}</tr></thead><tbody>${{bodyHtml}}</tbody></table></body></html>`;
  const blob = new Blob(['\ufeff' + doc], {{ type: 'application/vnd.ms-excel;charset=utf-8;' }});
  const url = URL.createObjectURL(blob);
  const link = document.createElement('a');
  link.href = url;
  link.download = 'ui_image_check_filtered.xls';
  document.body.appendChild(link);
  link.click();
  link.remove();
  URL.revokeObjectURL(url);
}}
function openPreview(button) {{
  const overlay = document.getElementById('overlay');
  document.getElementById('previewImage').src = button.dataset.fullSrc;
  document.getElementById('previewTitle').textContent = button.dataset.title || '';
  overlay.classList.add('open');
}}
function closePreview() {{
  document.getElementById('overlay').classList.remove('open');
  document.getElementById('previewImage').src = '';
}}
document.addEventListener('keydown', event => {{ if (event.key === 'Escape') closePreview(); }});
switchTab(activeTabType);
renderPage(1);
</script>
</body>
</html>
"""


def serve_report(
    report_path: Path,
    cache_path: Path,
    cache_db_path: Path | None = None,
    host: str = "127.0.0.1",
    port: int = 8765,
) -> int:
    report_path = report_path.resolve()
    cache_path = cache_path.resolve()
    cache_db_path = cache_db_path.resolve() if cache_db_path is not None else None
    root_dir = report_path.parent

    class ReportHandler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(root_dir), **kwargs)

        def do_POST(self) -> None:
            if self.path.split("?", 1)[0] != "/api/ocr-cache/operation":
                self.send_error(404)
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                data = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                if "relativePath" in data:
                    if cache_db_path is not None:
                        update_ocr_cache_operation_sqlite(
                            cache_db_path,
                            str(data["relativePath"]),
                            str(data["md5"]),
                            str(data["operation"]),
                        )
                    else:
                        update_ocr_cache_operation(
                            cache_path,
                            str(data["relativePath"]),
                            str(data["md5"]),
                            str(data["operation"]),
                        )
                else:
                    cache_path.write_text(
                        json.dumps(data, ensure_ascii=False, indent=2, default=str),
                        encoding="utf-8",
                    )
                response = b'{"ok":true}'
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(response)))
                self.end_headers()
                self.wfile.write(response)
            except Exception as exc:
                response = json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False).encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(response)))
                self.end_headers()
                self.wfile.write(response)

    class ReusableTCPServer(socketserver.TCPServer):
        allow_reuse_address = True

    with ReusableTCPServer((host, port), ReportHandler) as httpd:
        url = f"http://{host}:{httpd.server_address[1]}/{report_path.name}"
        print(f"报告服务已启动: {url}")
        print(f"OCR 缓存: {cache_db_path or cache_path}")
        try:
            os.startfile(url)
        except Exception:
            pass
        httpd.serve_forever()
    return 0


def add_thumbnail(
    ws,
    pil_image,
    xlsx_image,
    source: str,
    cell: str,
    thumb_dir: Path,
    temp_files: list[Path],
    no_resize: bool = False,
    max_box: tuple[int, int] | None = (96, 72),
) -> tuple[int, int, int]:
    local = local_image_path(source)
    if local is None:
        return 0, 0, 0
    try:
        with pil_image.open(local) as img:
            if not no_resize and max_box is not None:
                img.thumbnail(max_box)
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            thumb_path = thumb_dir / f"thumb_{cell}.png"
            img.save(thumb_path)
            out_w, out_h = img.size
        temp_files.append(thumb_path)
        drawing = xlsx_image(str(thumb_path))
        drawing.width = out_w
        drawing.height = out_h
        ws.add_image(drawing, cell)
        return 1, out_w, out_h
    except Exception as exc:
        if no_resize:
            label = "原图"
        elif max_box and max_box != (96, 72):
            label = f"{max(max_box)}px 限制图"
        else:
            label = "缩略图"
        print(f"WARN: 无法生成{label}，已保留路径文本: {source}: {exc}", file=sys.stderr)
        return 0, 0


def local_image_path(source: str) -> Path | None:
    if not source or source.lower().startswith(("http://", "https://")):
        return None
    path = Path(source)
    if path.exists() and path.is_file():
        return path
    return None


def _is_pillow_available() -> bool:
    try:
        from PIL import Image as _PILImage  # noqa: F401
        return True
    except ImportError:
        return False


def _require_pillow_for_report(output_path: Path) -> None:
    if _is_pillow_available():
        return
    raise SystemExit(
        "ERROR: 生成图片预览需要 Pillow，但当前 Python 环境没有安装 PIL。\n"
        f"当前 Python: {sys.executable}\n"
        f"请执行: \"{sys.executable}\" -m pip install Pillow\n"
        f"然后重新运行生成 {output_path.name}。"
    )


def parse_dt(value: str) -> dt.datetime:
    return dt.datetime.fromisoformat(value.replace("Z", "+00:00"))


def _join_config_path(root: str, path: str) -> str:
    path = path.replace("\\", "/").rstrip("/")
    if not root or path.lower().startswith(("http://", "https://")):
        return path
    if re.match(r"^[A-Za-z]:/", path) or path.startswith("/"):
        return path
    return f"{root.rstrip('/')}/{path}"


def _filter_whitelisted_files(
    files: dict[str, ImageFile],
    pair: dict[str, object],
) -> dict[str, ImageFile]:
    raw_whitelist = pair.get("whitelist", [])
    whitelist: list[str] = []
    whitelist_rules: list[dict[str, object]] = []
    for item in raw_whitelist:
        if isinstance(item, dict):
            whitelist_rules.append(item)
            continue
        parsed_rule = _parse_inline_whitelist_rule(str(item))
        if parsed_rule is not None:
            whitelist_rules.append(parsed_rule)
        else:
            whitelist.append(normalize_rel(str(item)))

    relative_roots = [
        normalize_rel(str(pair.get("mainland_relative", ""))),
        normalize_rel(str(pair.get("i18n_relative", ""))),
    ]
    ignored_prefixes: list[str] = []
    for item in whitelist:
        for root in relative_roots:
            if root and compare_key(item).startswith(compare_key(root) + "/"):
                ignored_prefixes.append(compare_key(item[len(root):].lstrip("/")))
                break
        else:
            ignored_prefixes.append(compare_key(item))

    return {
        key: value
        for key, value in files.items()
        if not any(key == prefix or key.startswith(prefix + "/") for prefix in ignored_prefixes)
        and not any(_matches_whitelist_rule(value.relative_path, rule, relative_roots) for rule in whitelist_rules)
    }


def _strip_pair_root(path: str, relative_roots: Sequence[str]) -> str:
    normalized = normalize_rel(path)
    for root in relative_roots:
        root = normalize_rel(root)
        if root and compare_key(normalized).startswith(compare_key(root) + "/"):
            return normalize_rel(normalized[len(root):].lstrip("/"))
    return normalized


def _parse_inline_whitelist_rule(item: str) -> dict[str, object] | None:
    normalized = normalize_rel(item)
    parts = normalized.split("/")
    for index, part in enumerate(parts):
        if part.startswith("^"):
            if index == 0:
                return None
            return {
                "directory_glob": "/".join(parts[:index]),
                "filename_regex": "/".join(parts[index:]),
            }
    return None


def _matches_whitelist_rule(
    relative_path: str,
    rule: dict[str, object],
    relative_roots: Sequence[str],
) -> bool:
    directory_glob = rule.get("directory_glob")
    filename_regex = rule.get("filename_regex")
    if not directory_glob or not filename_regex:
        return False

    path = _strip_pair_root(relative_path, relative_roots)
    path_key = compare_key(path)
    directory = str(PurePosixPath(path_key).parent)
    if directory == ".":
        directory = ""
    pattern = compare_key(_strip_pair_root(str(directory_glob), relative_roots))
    if not PurePosixPath(directory).match(pattern):
        return False

    filename = PurePosixPath(path).name
    stem = PurePosixPath(path).stem
    pattern_text = str(filename_regex)
    if re.fullmatch(pattern_text, filename, flags=re.IGNORECASE):
        return True
    if re.fullmatch(pattern_text, stem, flags=re.IGNORECASE):
        return True
    if pattern_text.endswith("$.*"):
        return re.fullmatch(pattern_text[:-2], stem, flags=re.IGNORECASE) is not None
    return False


def _load_config_pairs(config_path: str) -> list[dict[str, str]]:
    raw = Path(config_path).read_text(encoding="utf-8-sig")
    # tolerate Windows backslash paths in JSON
    raw = raw.replace("\\", "/")
    data = json.loads(raw)
    root = ""
    if isinstance(data, list):
        pairs = data
        whitelist = []
    elif isinstance(data, dict) and "pairs" in data:
        root = str(data.get("root", "")).replace("\\", "/").rstrip("/")
        whitelist = _normalize_whitelist(data.get("whitelist", []))
        pairs = data["pairs"]
    else:
        raise SystemExit(f"ERROR: 配置文件格式错误，需要 [{{\"name\":..., \"i18n\":..., \"mainland\":...}}, ...]")
    for p in pairs:
        if "i18n" not in p or "mainland" not in p:
            raise SystemExit(f"ERROR: 配置项缺少 i18n 或 mainland 字段: {p}")
        p["i18n_relative"] = p["i18n"].replace("\\", "/").strip("/")
        p["mainland_relative"] = p["mainland"].replace("\\", "/").strip("/")
        p["i18n"] = _join_config_path(root, p["i18n_relative"])
        p["mainland"] = _join_config_path(root, p["mainland_relative"])
        p["whitelist"] = _normalize_whitelist(p.get("whitelist", whitelist))
    return pairs


def _pair_report_type(pair: dict[str, str]) -> str:
    return (
        str(pair.get("type") or pair.get("category") or pair.get("name") or pair.get("mainland") or "")
        .strip()
    )


def _normalize_whitelist(items: object) -> list[object]:
    if not isinstance(items, list):
        return []
    normalized: list[object] = []
    for item in items:
        if isinstance(item, dict):
            rule = dict(item)
            if "directory_glob" in rule:
                rule["directory_glob"] = normalize_rel(str(rule["directory_glob"]))
            if "filename_regex" in rule:
                rule["filename_regex"] = str(rule["filename_regex"])
            normalized.append(rule)
        else:
            normalized.append(normalize_rel(str(item)))
    return normalized


def _finding_to_dict(f: Finding) -> dict:
    return {
        "category": f.category,
        "issue": f.issue,
        "relative_path": f.relative_path,
        "i18n_path": f.i18n_path,
        "mainland_path": f.mainland_path,
        "i18n_modified_at": f.i18n_modified_at.isoformat() if f.i18n_modified_at else None,
        "mainland_modified_at": f.mainland_modified_at.isoformat() if f.mainland_modified_at else None,
        "detail": f.detail,
        "mainland_created_at": f.mainland_created_at.isoformat() if f.mainland_created_at else None,
        "mainland_ocr_text": f.mainland_ocr_text,
        "i18n_ocr_text": f.i18n_ocr_text,
        "translation_status": f.translation_status,
        "translation_note": f.translation_note,
    }


def _save_intermediate(path: Path, findings: list[Finding], i18n_count: int, mainland_count: int) -> None:
    data = {
        "i18n_count": i18n_count,
        "mainland_count": mainland_count,
        "findings": [_finding_to_dict(f) for f in findings],
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def collect_findings(
    args: argparse.Namespace,
    report_callback: Callable[[list[Finding], dict[str, int]], None] | None = None,
) -> tuple[list[Finding], dict[str, int]]:
    last_check_at = args.since
    if args.assume_new_has_text:
        detector = assume_text_detector
    elif not args.no_ocr:
        if getattr(args, "ocr_cache_db", None):
            detector = sqlite_ocr_text_detector_factory(Path(args.ocr_cache_db))
        else:
            detector = ocr_text_detector_factory(cache_path=Path(args.ocr_cache_file))
    else:
        detector = default_text_detector

    if args.config:
        pairs = _load_config_pairs(args.config)
    else:
        pairs = [{"name": "", "i18n": args.i18n, "mainland": args.mainland}]

    all_findings: list[Finding] = []
    total_i18n = 0
    total_mainland = 0
    total_normal_synced = 0
    total_new_no_text = 0

    pair_count = len(pairs)

    for idx, pair in enumerate(pairs, 1):
        name = pair.get("name", "") or pair["mainland"]
        report_type = _pair_report_type(pair)
        log_step(f"开始处理配对目录 [{idx}/{pair_count}]: {name}")
        print(f"[{idx}/{pair_count}] 正在扫描: {name}", file=sys.stderr)
        is_remote = pair["i18n"].lower().startswith(("http://", "https://"))
        log_step(f"扫描国际版目录: {pair['i18n']}")
        print(f"  扫描国际版目录: {pair['i18n']}", file=sys.stderr)
        i18n_files = scan_svn(pair["i18n"]) if is_remote else scan_local(pair["i18n"], pair)
        log_step(f"扫描陆版目录: {pair['mainland']}")
        print(f"  扫描陆版目录: {pair['mainland']}", file=sys.stderr)
        mainland_files = scan_svn(pair["mainland"]) if is_remote else scan_local(pair["mainland"], pair)
        i18n_files = _filter_whitelisted_files(i18n_files, pair)
        mainland_files = _filter_whitelisted_files(mainland_files, pair)
        log_step(f"扫描完成: 国际版={len(i18n_files)} 陆版={len(mainland_files)}")
        print(f"  国际版 {len(i18n_files)} 张, 陆版 {len(mainland_files)} 张", file=sys.stderr)
        total_i18n += len(i18n_files)
        total_mainland += len(mainland_files)
        i18n_files, mainland_files = apply_max_file_sample(
            i18n_files, mainland_files, args.max_files,
        )
        log_step(
            f"应用采样后: max_files={args.max_files} 国际版={len(i18n_files)} 陆版={len(mainland_files)}"
        )
        log_step(f"开始比较目录: {name}")
        pair_findings, pair_stats = compare_category(
            report_type,
            i18n_files,
            mainland_files,
            last_check_at,
            detector,
            ocr_workers=args.ocr_workers,
        )
        all_findings.extend(pair_findings)
        total_normal_synced += pair_stats["normal_synced"]
        total_new_no_text += pair_stats["new_no_text"]
        log_step(
            f"目录比较完成: {name} normal_synced={pair_stats['normal_synced']} "
            f"new_no_text={pair_stats['new_no_text']} findings={len(pair_findings)}"
        )
        print(
            f"  正常同步 {pair_stats['normal_synced']} 张, "
            f"新增无文字 {pair_stats['new_no_text']} 张, "
            f"发现问题 {len(pair_findings)} 项, "
            f"累计 {len(all_findings)} 项",
            file=sys.stderr,
        )
        if report_callback is not None:
            report_callback(
                list(all_findings),
                {
                    "i18n_count": total_i18n,
                    "mainland_count": total_mainland,
                    "normal_synced": total_normal_synced,
                    "new_no_text": total_new_no_text,
                },
            )

    stats = {
        "i18n_count": total_i18n,
        "mainland_count": total_mainland,
        "normal_synced": total_normal_synced,
        "new_no_text": total_new_no_text,
    }
    return all_findings, stats


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="检查国际版/陆版图片差异，并输出 HTML。")
    parser.add_argument("--config", default=None, help="JSON 配置文件路径（默认自动查找当前目录 check_config.json）")
    parser.add_argument("--i18n", default=None, help="国际版图片目录路径（与 --mainland 配对，或使用 --config）")
    parser.add_argument("--mainland", default=None, help="陆版图片目录路径（与 --i18n 配对，或使用 --config）")
    parser.add_argument("--output", default="ui_image_check_report.html", help="输出文件路径，支持 .html / .htm")
    parser.add_argument(
        "--since",
        type=parse_dt,
        help="只检查此时间后的陆版新增图片，例如 2026-05-27T00:00:00+08:00",
    )
    parser.add_argument("--no-ocr", action="store_true", help="禁用 OCR 文字检测（默认启用）")
    parser.add_argument(
        "--ocr-workers",
        type=int,
        default=default_ocr_workers(),
        help="OCR 并发线程数，默认 1",
    )
    parser.add_argument(
        "--ocr-cache-file",
        default=OCR_CACHE_FILE,
        help=f"OCR 检测结果缓存文件（默认 {OCR_CACHE_FILE}），记录已检查文件避免重复 OCR",
    )
    parser.add_argument(
        "--ocr-cache-db",
        default=None,
        help="SQLite OCR 缓存数据库路径；指定后优先读写 SQLite，不再写 .ocr_cache.json",
    )
    parser.add_argument(
        "--max-files",
        type=int,
        help="测试用：每个对比目录按相对路径排序后最多取多少张图片",
    )
    parser.add_argument(
        "--assume-new-has-text",
        action="store_true",
        help="将陆版新增图片全部按有文字报告，用于无 OCR 环境人工复核",
    )
    parser.add_argument(
        "--max-image-px",
        type=int,
        default=None,
        help="HTML 报告预览图最长边像素上限（保持纵横比）；默认 720",
    )
    parser.add_argument(
        "--serve-report",
        action="store_true",
        help="生成报告后启动本地服务打开页面，使 HTML 可自动写回 OCR 缓存",
    )
    parser.add_argument(
        "--serve-port",
        type=int,
        default=8765,
        help="--serve-report 使用的本地端口，默认 8765",
    )
    args = parser.parse_args(argv)

    # auto-detect config file
    if not args.config and not args.i18n and not args.mainland:
        default_config = Path("check_config.json")
        if default_config.exists():
            args.config = str(default_config.resolve())

    if not args.config and (not args.i18n or not args.mainland):
        parser.error("请指定 --i18n 和 --mainland，或 --config 配置文件，或在当前目录放置 check_config.json")

    output_path = Path(args.output)
    if output_path.suffix.lower() not in {".html", ".htm"}:
        parser.error("输出文件只支持 .html / .htm")

    log_file, ocr_candidate_file = _init_run_outputs()
    log_step(format_version_line())
    log_step(f"开始检查: output={args.output}")
    log_step(f"日志文件: {log_file}")
    log_step(f"OCR 图片清单: {ocr_candidate_file}")
    _require_pillow_for_report(output_path)
    report_written = False

    def write_cumulative_report(current_findings: list[Finding], current_counts: dict[str, int]) -> None:
        nonlocal report_written
        log_step(
            f"生成累计 HTML 报告: {output_path} findings={len(current_findings)} "
            f"i18n={current_counts['i18n_count']} mainland={current_counts['mainland_count']}"
        )
        write_html_report(
            output_path,
            current_findings,
            i18n_count=current_counts["i18n_count"],
            mainland_count=current_counts["mainland_count"],
            normal_synced=current_counts["normal_synced"],
            new_no_text=current_counts["new_no_text"],
            max_image_px=args.max_image_px,
            ocr_cache_name=Path(args.ocr_cache_db).name if args.ocr_cache_db else Path(args.ocr_cache_file).name,
        )
        report_written = True

    findings, counts = collect_findings(args, report_callback=write_cumulative_report)
    _write_ocr_candidate_list()
    log_step(f"OCR 图片清单已写入: {ocr_candidate_file} count={len(_OCR_CANDIDATES)}")
    log_new_no_text_findings(findings)
    if not report_written:
        log_step(f"开始生成 HTML 报告: {output_path}")
        write_cumulative_report(findings, counts)
    log_step(f"检查完成: findings={len(findings)} output={args.output}")
    print(f"检查完成: {len(findings)} 项，输出: {args.output}")
    if args.serve_report:
        cache_db_path = Path(args.ocr_cache_db) if args.ocr_cache_db else None
        return serve_report(
            output_path,
            Path(args.ocr_cache_file),
            cache_db_path=cache_db_path,
            port=args.serve_port,
        )
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
