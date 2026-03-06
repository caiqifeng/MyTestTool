"""
批量对比 UI 目录下的大图标和小图标，生成带图片的 Excel 表格
使用项目配置文件来匹配图标对
"""

import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pic_compare import compare_images
from PIL import Image
from datetime import datetime
import io


LogFullPath = "Log_" + datetime.now().strftime("%m%d%H%M%S") + ".txt"
f = open(LogFullPath, "a", encoding="utf-8")
def write_log(msg):
    print(msg)
    f.write(msg + "\n")
    f.flush()

def get_image_info(dir_path):
    if not dir_path.exists():
        return "不存在"

    try:
        with Image.open(dir_path) as img:
            image_info = str(img.width) + "x" + str(img.height)
            return image_info

    except Exception as e:
        #write_log(f"错误: {e}")
        return "获取异常"

def parse_config_file(config_file_path):
    """
    解析项目配置文件，提取UIBigIcon和UIIcon的路径对
    
    Args:
        config_file_path: 配置文件路径（UIIcon.txt）
    
    Returns:
        list: [(icon_name, bigicon_path, icon_path), ...]
    """
    matches = []
    
    with open(config_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 正则匹配UIBigIcon和UIIcon的路径
    # 匹配两种格式：
    # 1. "Assets/_Game/Resource/ArtSource/UI/..."
    # 2. "ArtSource://UI/..."
    pattern = r'"UIBigIcon":\s*"([^"]+)"\s*,\s*[\r\n]+\s*[^\r\n]*"UIIcon":\s*"([^"]+)"'
    
    found_pairs = re.findall(pattern, content, re.MULTILINE)
    
    for big_icon_raw, icon_raw in found_pairs:
        # 跳过空路径
        if not big_icon_raw or not icon_raw:
            continue
        
        # 清理路径前缀
        def clean_path(path):
            # 移除 "Assets/_Game/Resource/ArtSource/UI/"
            path = re.sub(r'^Assets/_Game/Resource/ArtSource/UI/', '', path)
            # 移除 "ArtSource://UI/"
            path = re.sub(r'^ArtSource://UI/', '', path)
            return path
        
        big_icon_path = clean_path(big_icon_raw)
        icon_path = clean_path(icon_raw)
        
        # 提取图标名称（使用文件名不含扩展名）
        icon_name = Path(icon_path).stem
        
        matches.append((icon_name, big_icon_path, icon_path))
    
    return matches


def index_png_files(directory):
    """
    递归索引目录下所有的 PNG 文件
    
    Args:
        directory: 要索引的目录路径
    
    Returns:
        dict: {文件名: 完整路径} 的字典
    """
    png_files = {}
    directory_path = Path(directory)
    
    if not directory_path.exists():
        write_log(f"警告: 目录不存在 - {directory}")
        return png_files
    
    # 递归查找所有 .png 文件
    for png_file in directory_path.rglob("*.png"):
        filename = png_file.name
        # 如果有重名文件，保存相对路径作为key
        if filename in png_files:
            # 使用相对于根目录的路径作为key
            rel_path = png_file.relative_to(directory_path)
            key = str(rel_path).replace("\\", "/")
        else:
            key = filename
        
        png_files[key] = str(png_file)
    
    return png_files


def match_icons(bigicon_files, icon_files):
    """
    根据文件名匹配大图标和小图标
    
    Args:
        bigicon_files: 大图标文件字典
        icon_files: 小图标文件字典
    
    Returns:
        list: [(图标名称, 大图路径, 小图路径), ...]
    """
    matches = []
    
    # 为小图标创建一个快速查找字典（支持文件名和相对路径）
    icon_lookup = {}
    for key, path in icon_files.items():
        # 同时用文件名和相对路径作为key
        filename = Path(key).name
        icon_lookup[filename] = path
        icon_lookup[key] = path
    
    # 遍历大图标，查找对应的小图标
    for big_key, big_path in bigicon_files.items():
        big_filename = Path(big_key).name
        
        # 尝试多种匹配方式
        small_path = None
        
        # 1. 先尝试完全匹配文件名
        if big_filename in icon_lookup:
            small_path = icon_lookup[big_filename]
        
        # 2. 尝试匹配相对路径（去掉 BigIcon 前缀）
        elif big_key in icon_lookup:
            small_path = icon_lookup[big_key]
        
        # 3. 尝试匹配去掉 "BigIcon/" 后的路径
        if not small_path and isinstance(big_key, str):
            # 从相对路径中提取子路径
            big_path_obj = Path(big_path)
            # 找到 BigIcon 后面的路径部分
            parts = big_path_obj.parts
            try:
                bigicon_idx = parts.index("BigIcon")
                sub_path = Path(*parts[bigicon_idx + 1:])
                sub_path_str = str(sub_path).replace("\\", "/")
                
                # 在 Icon 目录下查找相同的子路径
                if sub_path_str in icon_lookup:
                    small_path = icon_lookup[sub_path_str]
                elif sub_path.name in icon_lookup:
                    small_path = icon_lookup[sub_path.name]
            except (ValueError, IndexError):
                pass
        
        if small_path:
            # 提取图标名称（不含扩展名）
            icon_name = Path(big_filename).stem
            matches.append((icon_name, big_path, small_path))
        else:
            write_log(f"未找到匹配: {big_filename}")
    
    return matches


def resize_image_for_excel(image_path, max_width=100, max_height=100):
    """
    调整图片大小以适应 Excel 单元格
    
    Args:
        image_path: 图片路径
        max_width: 最大宽度（像素）
        max_height: 最大高度（像素）
    
    Returns:
        BytesIO: 调整后的图片数据流
    """
    try:
        img = Image.open(image_path)
        
        # 保持纵横比调整大小
        img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
        
        # 转换为 RGB（如果是 RGBA）
        if img.mode == 'RGBA':
            # 创建白色背景
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])  # 使用 alpha 通道作为蒙版
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # 保存到内存
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return img_buffer
    except Exception as e:
        write_log(f"  警告: 无法处理图片 {image_path}: {e}")
        return None


def compare_all_icons_to_excel(ui_directory, config_file, output_file="comparison_results.xlsx"):
    """
    根据配置文件对比所有匹配的图标并生成带图片的 Excel 报告
    
    Args:
        ui_directory: UI 根目录路径
        config_file: 配置文件路径（UIIcon.txt）
        output_file: 输出 Excel 文件名
    """
    ui_path = Path(ui_directory)
    
    write_log("正在解析配置文件...")

    matches = parse_config_file(config_file)
    write_log(f"从配置文件中找到 {len(matches)} 对图标配置\n")
    
    if not matches:
        write_log("配置文件中没有找到有效的图标对，程序退出")
        return
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图标相似度对比"
    
    # 设置表头
    headers = ['图片名', '大图路径', '小图路径', '大图图片', '小图图片', '感知相似度', 'SSIM', '像素相似度']
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 30  # 图片名
    ws.column_dimensions['B'].width = 40  # 大图路径
    ws.column_dimensions['C'].width = 40  # 小图路径
    ws.column_dimensions['D'].width = 15  # 大图图片
    ws.column_dimensions['E'].width = 15  # 小图图片
    ws.column_dimensions['F'].width = 15  # 感知相似度
    ws.column_dimensions['G'].width = 12  # SSIM
    ws.column_dimensions['H'].width = 15  # 像素相似度
    
    # 设置行高（标准行高适合 100px 的图片）
    ws.row_dimensions[1].height = 30  # 表头
    
    total = len(matches)
    write_log("="*80)
    write_log(f"开始对比 {total} 对图标并生成 Excel...")
    write_log("="*80 + "\n")
    
    current_row = 2
    success_count = 0
    error_count = 0
    skip_count = 0
    
    for idx, (icon_name, big_rel_path, small_rel_path) in enumerate(matches, 1):
        # 构建完整路径
        big_path = ui_path / big_rel_path
        small_path = ui_path / small_rel_path
        
        big_image_info = get_image_info(big_path)
        small_image_info = get_image_info(small_path)

        
        write_log(f"[{idx}/{total}] 处理: {icon_name}")
        write_log(f"  大图: {big_rel_path} \t {big_image_info}")
        write_log(f"  小图: {small_rel_path} \t {small_image_info}")
        
        # 检查文件是否存在
        if not big_path.exists():
            write_log(f"  警告: 大图不存在，跳过")
            skip_count += 1
            write_log(f"\n")
            continue
        
        if not small_path.exists():
            write_log(f"  警告: 小图不存在，跳过")
            skip_count += 1
            write_log(f"\n")
            continue
        
        try:
            # 计算相似度
            comparison = compare_images(str(big_path), str(small_path), show_details=False)
            perc_sim = comparison['Perceptual_Similarity']
            ssim_score = comparison['SSIM']
            pixel_sim = comparison['Pixel_Similarity']
            
            # 写入文本数据
            ws.cell(row=current_row, column=1, value=icon_name)
            ws.cell(row=current_row, column=2, value=big_rel_path)
            ws.cell(row=current_row, column=3, value=small_rel_path)
            ws.cell(row=current_row, column=6, value=f"{perc_sim:.4f}")
            ws.cell(row=current_row, column=7, value=f"{ssim_score:.4f}")
            ws.cell(row=current_row, column=8, value=f"{pixel_sim:.4f}")
            
            # 设置文本对齐
            for col_num in range(1, 9):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 相似度列居中
            for col_num in [6, 7, 8]:
                ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
            
            # 插入大图
            big_img_buffer = resize_image_for_excel(str(big_path), max_width=100, max_height=100)
            if big_img_buffer:
                xl_big_img = XLImage(big_img_buffer)
                # 设置图片大小
                xl_big_img.width = 100
                xl_big_img.height = 100
                # 将图片锚定到单元格 D (列4)
                ws.add_image(xl_big_img, f'D{current_row}')
            
            # 插入小图
            small_img_buffer = resize_image_for_excel(str(small_path), max_width=100, max_height=100)
            if small_img_buffer:
                xl_small_img = XLImage(small_img_buffer)
                # 设置图片大小
                xl_small_img.width = 100
                xl_small_img.height = 100
                # 将图片锚定到单元格 E (列5)
                ws.add_image(xl_small_img, f'E{current_row}')
            
            # 设置行高以容纳图片（约 100px + 边距）
            ws.row_dimensions[current_row].height = 80
            
            write_log(f"  感知相似度: {perc_sim:.4f}, SSIM: {ssim_score:.4f} ✓")
            
            current_row += 1
            success_count += 1
            
        except Exception as e:
            write_log(f"  错误: {e}")
            # 即使出错也记录
            ws.cell(row=current_row, column=1, value=icon_name)
            ws.cell(row=current_row, column=2, value=big_rel_path)
            ws.cell(row=current_row, column=3, value=small_rel_path)
            ws.cell(row=current_row, column=6, value="ERROR")
            ws.cell(row=current_row, column=7, value="ERROR")
            ws.cell(row=current_row, column=8, value="ERROR")
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            error_count += 1
        
        write_log(f"\n")
    
    # 保存 Excel 文件
    output_path = Path(output_file)
    write_log("="*80)
    write_log(f"保存结果到: {output_path.absolute()}")
    wb.save(output_path)
    write_log("="*80 + "\n")
    
    write_log(f"✓ 成功保存 Excel 文件")
    write_log(f"  成功: {success_count}")
    write_log(f"  失败: {error_count}")
    write_log(f"  跳过: {skip_count}")
    write_log(f"  总计: {len(matches)}")
    write_log("="*80)


def main():
    """
    主函数
    """
    import argparse
    
    parser = argparse.ArgumentParser(
        description='批量对比 UI 目录下的大图标和小图标，生成带图片的 Excel 表格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python batch_compare_excel.py UI UIIcon.txt
  python batch_compare_excel.py UI UIIcon.txt -o my_results.xlsx
        """
    )
    
    parser.add_argument('ui_directory', 
                       help='UI 根目录路径')
    parser.add_argument('config_file', 
                       help='配置文件路径 (UIIcon.txt)')
    parser.add_argument('-o', '--output', default='icon_comparison.xlsx',
                       help='输出 Excel 文件名 (默认: icon_comparison.xlsx)')
    
    args = parser.parse_args()
    
    try:
        compare_all_icons_to_excel(args.ui_directory, args.config_file, args.output)
        return 0
    except Exception as e:
        write_log(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
