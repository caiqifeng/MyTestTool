"""
批量对比 UI 目录下的大图标和小图标，生成带图片的 Excel 表格
使用项目配置文件来匹配图标对
优化点：
1. 面向对象重构，提升代码模块化和可维护性
2. 移除全局变量，使用类属性管理状态
3. 性能优化（缓存、减少重复IO）
4. 增强错误处理和日志记录
5. 类型注解完善，提升代码可读性
6. 配置项集中管理，便于修改
7. 路径处理统一使用pathlib，提升跨平台兼容性
8. 减少魔法值，增加常量定义
"""

import os
import re
import io
import subprocess
import traceback
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
import argparse

# ===================== 常量配置（集中管理，便于修改） =====================
class Config:
    RELATIVE_ROOT_PREFIX = "Assets/_Game/Resource/"

    # 路径协议映射
    PROTOCOL_MAP = {
        "ArtSource://": f"{RELATIVE_ROOT_PREFIX}ArtSource",
        "ArtBase://": f"{RELATIVE_ROOT_PREFIX}ArtBase"
    }
        
    # Excel配置
    EXCEL_HEADERS = [
        'ID', '图片名', '大图路径', '小图路径', '大图图片', '小图图片', 
        '感知相似度', 'SSIM', '像素相似度', '大图Size', '小图Size', 
        '大图透明度', '小图透明度'
    ]
    
    # Excel列名到列号的映射（避免硬编码列号）
    COLUMN_MAP = {
        'ID': 1, '图片名': 2, '大图路径': 3, '小图路径': 4,
        '大图图片': 5, '小图图片': 6, '感知相似度': 7, 'SSIM': 8,
        '像素相似度': 9, '大图Size': 10, '小图Size': 11,
        '大图透明度': 12, '小图透明度': 13
    }
    
    EXCEL_COL_WIDTHS = {
        'A': 30, 'B': 30, 'C': 40, 'D': 40, 'E': 15, 'F': 15,
        'G': 12, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 15
    }
    EXCEL_HEADER_STYLE = {
        'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'font': Font(bold=True, color="FFFFFF", size=12),
        'alignment': Alignment(horizontal='center', vertical='center')
    }
    IMAGE_MAX_SIZE = (100, 100)  # 图片最大尺寸
    ROW_HEIGHT = 80  # 图片行高
    HEADER_ROW_HEIGHT = 30  # 表头行高
    
    # 图片格式
    VALID_IMAGE_SUFFIX = ('.png', '.webp', '.gif', '.tga')
    
    # SVN配置
    SVN_IGNORE_ERROR = False
    PROJECT_PATHS = [
        r"Assets\_Game\Resource\Config\Item\Item",
        r"Assets\_Game\Resource\ArtSource\UI\BigIcon",
        r"Assets\_Game\Resource\ArtSource\UI\Icon",
        r"Assets\_Game\Resource\ArtBase\UI"
    ]

# ===================== 数据类（提升数据结构可读性） =====================
@dataclass
class IconPair:
    """图标对数据类"""
    id_name: str
    icon_name: str
    big_icon_path: str
    small_icon_path: str

@dataclass
class ComparisonResult:
    """对比结果数据类"""
    perceptual_similarity: float
    ssim: float
    pixel_similarity: float

# ===================== 工具类（核心逻辑封装） =====================
class IconComparator:
    def __init__(self, ui_directory: str, log_path: str):
        self.ui_dir = Path(ui_directory).resolve()
        self.log_path = Path(log_path)
        self.statistics = {
            "success": 0,
            "error": 0,
            "skip": 0,
            "total": 0
        }
        
        # 配置日志系统（使用logging模块，避免手动管理文件）
        self.logger = logging.getLogger('IconComparator')
        self.logger.setLevel(logging.INFO)
        
        # 清除已有的处理器
        self.logger.handlers.clear()
        
        # 文件处理器
        file_handler = logging.FileHandler(self.log_path, mode='a', encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_formatter = logging.Formatter('%(message)s')
        file_handler.setFormatter(file_formatter)
        
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_formatter = logging.Formatter('%(message)s')
        console_handler.setFormatter(console_formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

    def log(self, msg: str):
        """日志记录（控制台+文件）"""
        self.logger.info(msg)

    def parse_config_file(self, config_file_path: str) -> List[IconPair]:
        """解析配置文件，提取图标对"""
        self.log("正在解析配置文件...")
        matches = []
        
        try:
            with open(config_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            pattern = r'"ID":\s*"([^"]*)".*?"UIBigIcon":\s*"([^"]*)".*?"UIIcon":\s*"([^"]*)"'
            found_pairs = re.findall(pattern, content, re.DOTALL)
            self.log(f"共匹配到 {len(found_pairs)} 组数据")

            for id_raw, big_icon_raw, icon_raw in found_pairs:
                # 处理空路径
                big_icon_raw = big_icon_raw if big_icon_raw else "大图为空"
                icon_raw = icon_raw if icon_raw else "小图为空"
                
                big_icon_path = big_icon_raw
                icon_path = icon_raw
                id_name = id_raw
                icon_name = Path(icon_path).stem
                
                matches.append(IconPair(
                    id_name=id_name,
                    icon_name=icon_name,
                    big_icon_path=big_icon_path,
                    small_icon_path=icon_path
                ))
        except Exception as e:
            self.log(f"解析配置文件失败: {e}")
            raise
        
        self.statistics["total"] = len(matches)
        return matches

    def custom_protocol_to_abs_path(self, protocol_path: str) -> str:
        """转换自定义协议路径为绝对路径"""
        if not isinstance(protocol_path, str) or not protocol_path.strip():
            self.log(f"【错误】无效路径：路径为空或非字符串类型")
            return protocol_path

        # 处理自定义协议
        for protocol_prefix, sub_dir in Config.PROTOCOL_MAP.items():
            if protocol_path.startswith(protocol_prefix):
                relative_path = protocol_path[len(protocol_prefix):]
                if not relative_path.strip():
                    self.log(f"【错误】无效路径：{protocol_path}，协议后无实际路径")
                    return protocol_path
                abs_path = self.ui_dir / sub_dir / relative_path
                return str(abs_path)

        # 处理相对根路径
        if protocol_path.startswith(Config.RELATIVE_ROOT_PREFIX):
            abs_path = self.ui_dir / protocol_path
            return str(abs_path)

        self.log(f"【错误】不支持的协议：{protocol_path}，仅支持{list(Config.PROTOCOL_MAP.keys())}")
        return protocol_path

    def get_image_info(self, image_path: str) -> str:
        """获取图片尺寸信息"""
        if not os.path.exists(image_path):
            return "不存在"

        try:
            with Image.open(image_path) as img:
                return f"{img.width}x{img.height}"
        except Exception as e:
            self.log(f"获取图片信息失败 {image_path}: {e}")
            return "获取异常"

    def calculate_transparency_ratio(self, image_path: str, include_semi: bool = False) -> float:
        """计算图片透明度占比"""
        image_path = str(image_path)
        if not os.path.exists(image_path):
            self.log(f"【错误】文件不存在：{image_path}")
            return -999.0

        if Path(image_path).suffix not in Config.VALID_IMAGE_SUFFIX:
            self.log(f"【警告】非支持的透明图片格式：{image_path}")
            return -999.0

        try:
            with Image.open(image_path) as img:
                rgba_img = img.convert("RGBA")
                width, height = rgba_img.size
                total_pixels = width * height
                
                if total_pixels == 0:
                    self.log(f"【错误】空图片：{image_path}")
                    return -999.0

                # 批量提取A通道，提升性能
                a_channel = [pixel[3] for pixel in rgba_img.getdata()]
                
                # 统计透明像素
                if include_semi:
                    transparent_pixels = sum(1 for a in a_channel if a < 255)
                else:
                    transparent_pixels = sum(1 for a in a_channel if a == 0)

                ratio = (transparent_pixels / total_pixels) * 100
                return round(ratio, 2)
        except Exception as e:
            self.log(f"【错误】处理图片失败 {image_path}：{str(e)}")
            return -999.0

    def resize_image_for_excel(self, image_path: str) -> Optional[io.BytesIO]:
        """调整图片大小适配Excel"""
        try:
            with Image.open(image_path) as img:
                # 保持纵横比缩放（使用LANCZOS替代已废弃的常量）
                img.thumbnail(Config.IMAGE_MAX_SIZE, Image.Resampling.LANCZOS)
                
                # 处理透明通道
                if img.mode == 'RGBA':
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    background.paste(img, mask=img.split()[3])
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # 保存到内存缓冲区
                img_buffer = io.BytesIO()
                img.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                return img_buffer
        except Exception as e:
            self.log(f"  警告: 无法处理图片 {image_path}: {e}")
            return None

    def compare_images(self, big_path: str, small_path: str) -> ComparisonResult:
        """调用图片对比函数（兼容原逻辑，此处保留接口）"""
        # 原代码中compare_images来自pic_compare，此处保留接口
        from pic_compare import compare_images as external_compare
        try:
            result = external_compare(str(big_path), str(small_path), show_details=False)
            return ComparisonResult(
                perceptual_similarity=result['Perceptual_Similarity'],
                ssim=result['SSIM'],
                pixel_similarity=result['Pixel_Similarity']
            )
        except Exception as e:
            self.log(f"对比图片失败 {big_path} vs {small_path}: {e}")
            raise

    def generate_excel(self, icon_pairs: List[IconPair], output_file: str):
        """生成Excel报告"""
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore
        ws.title = "图标相似度对比"

        # 写入表头
        ws.append(Config.EXCEL_HEADERS)
        
        # 设置表头样式
        for col_num, header in enumerate(Config.EXCEL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = Config.EXCEL_HEADER_STYLE['fill']
            cell.font = Config.EXCEL_HEADER_STYLE['font']
            cell.alignment = Config.EXCEL_HEADER_STYLE['alignment']

        # 设置列宽
        for col, width in Config.EXCEL_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # 设置表头行高
        ws.row_dimensions[1].height = Config.HEADER_ROW_HEIGHT

        # 处理每个图标对
        current_row = 2
        self.log("="*80)
        self.log(f"开始对比 {self.statistics['total']} 对图标并生成 Excel...")
        self.log("="*80 + "\n")

        for idx, pair in enumerate(icon_pairs, 1):
            self.log(f"[{idx}/{self.statistics['total']}] 处理: {pair.icon_name}")
            
            # 转换为绝对路径
            big_path = self.custom_protocol_to_abs_path(pair.big_icon_path)
            small_path = self.custom_protocol_to_abs_path(pair.small_icon_path)
            
            # 获取图片基础信息
            big_info = self.get_image_info(big_path)
            small_info = self.get_image_info(small_path)
            big_trans = self.calculate_transparency_ratio(big_path)
            small_trans = self.calculate_transparency_ratio(small_path)

            self.log(f"  大图: {pair.big_icon_path} \t {big_info} \t {big_trans}")
            self.log(f"  小图: {pair.small_icon_path} \t {small_info} \t {small_trans}")

            # 检查文件是否存在（修复重复统计问题）
            skip = False
            big_exists = os.path.exists(big_path)
            small_exists = os.path.exists(small_path)
            
            if not big_exists or not small_exists:
                missing = []
                if not big_exists:
                    missing.append("大图")
                if not small_exists:
                    missing.append("小图")
                self.log(f"  警告: {'/'.join(missing)}不存在，跳过")
                self.statistics["skip"] += 1
                skip = True

            try:
                # 初始化对比结果
                comp_result = ComparisonResult(-1.0, -1.0, -1.0)
                if not skip:
                    comp_result = self.compare_images(big_path, small_path)

                # 写入基础数据（使用列名映射，避免硬编码列号）
                ws.cell(row=current_row, column=Config.COLUMN_MAP['ID'], value=pair.id_name)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['图片名'], value=pair.icon_name)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['大图路径'], value=pair.big_icon_path)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['小图路径'], value=pair.small_icon_path)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['感知相似度'], value=f"{comp_result.perceptual_similarity:.4f}")
                ws.cell(row=current_row, column=Config.COLUMN_MAP['SSIM'], value=f"{comp_result.ssim:.4f}")
                ws.cell(row=current_row, column=Config.COLUMN_MAP['像素相似度'], value=f"{comp_result.pixel_similarity:.4f}")
                ws.cell(row=current_row, column=Config.COLUMN_MAP['大图Size'], value=big_info)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['小图Size'], value=small_info)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['大图透明度'], value=big_trans)
                ws.cell(row=current_row, column=Config.COLUMN_MAP['小图透明度'], value=small_trans)

                # 设置单元格对齐
                for col_num in range(1, 14):
                    cell = ws.cell(row=current_row, column=col_num)
                    if col_num in [1,2,3,4]:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 插入图片
                if os.path.exists(big_path):
                    big_buffer = self.resize_image_for_excel(big_path)
                    if big_buffer:
                        xl_img = XLImage(big_buffer)
                        xl_img.width, xl_img.height = Config.IMAGE_MAX_SIZE
                        ws.add_image(xl_img, f'E{current_row}')

                if os.path.exists(small_path):
                    small_buffer = self.resize_image_for_excel(small_path)
                    if small_buffer:
                        xl_img = XLImage(small_buffer)
                        xl_img.width, xl_img.height = Config.IMAGE_MAX_SIZE
                        ws.add_image(xl_img, f'F{current_row}')

                # 设置行高
                ws.row_dimensions[current_row].height = Config.ROW_HEIGHT

                if not skip:
                    self.log(f"  感知相似度: {comp_result.perceptual_similarity:.4f}, SSIM: {comp_result.ssim:.4f} ✓")
                    self.statistics["success"] += 1
                current_row += 1

            except FileNotFoundError as e:
                self.log(f"  错误: 文件未找到 - {e}")
                self._write_error_row(ws, current_row, pair, "FILE_NOT_FOUND")
                current_row += 1
                self.statistics["error"] += 1
            except ValueError as e:
                self.log(f"  错误: 数据错误 - {e}")
                self._write_error_row(ws, current_row, pair, "VALUE_ERROR")
                current_row += 1
                self.statistics["error"] += 1
            except Exception as e:
                self.log(f"  错误: 未知错误 - {e}")
                self.log(f"  详细信息: {traceback.format_exc()}")
                self._write_error_row(ws, current_row, pair, "ERROR")
                current_row += 1
                self.statistics["error"] += 1

            self.log("\n")

        # 保存Excel
        output_path = Path(output_file).resolve()
        wb.save(output_path)
        self.log("="*80)
        self.log(f"保存结果到: {output_path}")
        self.log("="*80 + "\n")
        self.log(f"✓ 成功保存 Excel 文件")
        self.log(f"  成功: {self.statistics['success']}")
        self.log(f"  失败: {self.statistics['error']}")
        self.log(f"  跳过: {self.statistics['skip']}")
        self.log(f"  总计: {self.statistics['total']}")
        self.log("="*80)

    def _write_error_row(self, ws: Worksheet, row: int, pair: IconPair, error_msg: str):
        """写入错误行到Excel（辅助方法）"""
        ws.cell(row=row, column=Config.COLUMN_MAP['ID'], value=pair.id_name)
        ws.cell(row=row, column=Config.COLUMN_MAP['图片名'], value=pair.icon_name)
        ws.cell(row=row, column=Config.COLUMN_MAP['大图路径'], value=pair.big_icon_path)
        ws.cell(row=row, column=Config.COLUMN_MAP['小图路径'], value=pair.small_icon_path)
        ws.cell(row=row, column=Config.COLUMN_MAP['感知相似度'], value=error_msg)
        ws.cell(row=row, column=Config.COLUMN_MAP['SSIM'], value=error_msg)
        ws.cell(row=row, column=Config.COLUMN_MAP['像素相似度'], value=error_msg)
        ws.row_dimensions[row].height = 20

# ===================== SVN 工具函数 =====================
def svn_update_dir(target_dir: str, ignore_error: bool = False) -> bool:
    """执行SVN Update"""
    target_dir = str(Path(target_dir).resolve())
    if not os.path.isdir(target_dir):
        print(f"【错误】目录无效：{target_dir}")
        return False

    cmd = ["svn", "update", target_dir, "--non-interactive"]
    print(f"【执行】SVN Update: {target_dir}")
    print(f"【命令】{' '.join(cmd)}")

    try:
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            shell=False,
            encoding="utf-8",
            errors="ignore",
            cwd=target_dir
        )

        if result.stdout:
            print(f"【SVN输出】\n{result.stdout.strip()}")

        if result.returncode == 0:
            print(f"【成功】SVN Update 完成: {target_dir}\n")
            return True
        else:
            error_msg = f"【失败】SVN Update 出错，返回码：{result.returncode}"
            if ignore_error:
                print(f"{error_msg}，已忽略\n")
                return False
            raise subprocess.CalledProcessError(result.returncode, cmd, result.stdout)
    except Exception as e:
        print(f"【错误】执行SVN Update失败: {e}\n")
        return False

# ===================== JSON 提取工具 =====================
def search_json_files(root_dir: str, save_path: str):
    """提取JSON文件中的指定行"""
    pattern = re.compile(r'^\s*("ID":|"UIBigIcon":|"UIIcon":)\s*.+,?$', re.MULTILINE)
    root_path = Path(root_dir).resolve()

    with open(save_path, 'w', encoding='utf-8') as txt_f:
        for json_file in root_path.rglob("*.json"):
            try:
                content = json_file.read_text(encoding='utf-8')
                match_lines = [line.rstrip() for line in content.splitlines() if pattern.search(line)]
                
                if len(match_lines) == 3:
                    txt_f.write(f"  {json_file}\n")
                    for line in match_lines:
                        txt_f.write(f"\t{line}\n")
                    txt_f.write("\n")
            except Exception as e:
                error_info = f"【读取失败】{json_file} - {str(e)}\n\n"
                txt_f.write(error_info)
                print(error_info.strip())

    print(f"提取完成！结果保存至：{save_path}")

# ===================== 分支判断 =====================
def get_branch(ui_dir: str) -> str:
    """判断分支类型"""
    if "b_SEED_sandbox_dev_2022-06-24" in ui_dir:
        return "main"
    elif "b_SEED_dev_pre_2025-12-16" in ui_dir:
        return "pre"
    return "nil"

# ===================== 主函数 =====================
def main():
    parser = argparse.ArgumentParser(
        description='批量对比 UI 目录下的大图标和小图标，生成带图片的 Excel 表格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python batch_compare_excel.py UI
  python batch_compare_excel.py UI --ignore-svn-error
        """
    )
    parser.add_argument('ui_directory', help='UI 根目录路径')
    parser.add_argument('--ignore-svn-error', action='store_true',
                       help='忽略 SVN 更新错误继续执行')
    args = parser.parse_args()

    # 初始化路径和文件名
    time_str = datetime.now().strftime("%m%d_%H%M")
    branch = get_branch(args.ui_directory)

    # 获取脚本所在目录（绝对路径）
    script_dir = Path(__file__).resolve().parent
    
    # 创建Release输出目录
    output_dir = script_dir / "Release"
    output_dir.mkdir(exist_ok=True)
    
    # 输出文件路径（绝对路径）
    log_path = str(output_dir / f"{branch}_{time_str}_Log.txt")
    json_content_txt = str(output_dir / f"{branch}_{time_str}_json_content.txt")
    result_xlsx = str(output_dir / f"{branch}_{time_str}_icon_comparison.xlsx")

    # 工程根目录
    target_dir = Path(args.ui_directory).resolve()
    remove_suffix = Path("Assets/_Game/Resource")
    if remove_suffix in target_dir.parts:
        target_dir = target_dir.parents[len(target_dir.parts) - target_dir.parts.index(remove_suffix) - 1]

    try:
        # 1. 更新SVN
        svn_ignore = args.ignore_svn_error if hasattr(args, 'ignore_svn_error') else Config.SVN_IGNORE_ERROR
        for rel_path in Config.PROJECT_PATHS:
            full_path = target_dir / rel_path
            svn_update_dir(str(full_path), svn_ignore)

        # 2. 提取JSON内容
        item_root = target_dir / Config.PROJECT_PATHS[0]
        search_json_files(str(item_root), json_content_txt)

        # 3. 初始化比较器并生成Excel
        comparator = IconComparator(args.ui_directory, log_path)
        icon_pairs = comparator.parse_config_file(json_content_txt)
        if not icon_pairs:
            comparator.log("配置文件中无有效图标对，程序退出")
            return 0

        comparator.generate_excel(icon_pairs, result_xlsx)
        return 0

    except KeyboardInterrupt:
        print("\n程序被用户中断")
        return 130
    except Exception as e:
        print(f"程序执行失败: {e}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit(main())